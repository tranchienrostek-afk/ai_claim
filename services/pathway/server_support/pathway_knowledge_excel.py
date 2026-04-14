from __future__ import annotations

import json
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.worksheet.table import Table, TableStyleInfo

from server_support.paths import BASE_DIR, KNOWLEDGE_EXCEL_VIEWS_DIR


class PathwayKnowledgeExcelBridge:
    """Export ontology-backed knowledge assets to workbook views and sync edits back into Neo4j."""

    def __init__(self, ontology_store, knowledge_registry_store, knowledge_text_bridge=None):
        self.ontology_store = ontology_store
        self.knowledge_registry_store = knowledge_registry_store
        self.knowledge_text_bridge = knowledge_text_bridge
        KNOWLEDGE_EXCEL_VIEWS_DIR.mkdir(parents=True, exist_ok=True)

    def export_asset_workbook(self, asset_id: str) -> dict[str, Any]:
        asset = self.knowledge_registry_store.get_asset(asset_id)
        if asset is None:
            raise KeyError(f"Asset not found: {asset_id}")

        graph_trace = self.knowledge_registry_store.graph_trace(asset_id)
        doc_ids = [row.get("doc_id") for row in graph_trace.get("ontology_documents") or [] if row.get("doc_id")]
        if not doc_ids:
            raise ValueError("Asset has no ontology-backed document in graph yet")

        payload = self._collect_asset_payload(asset_id, doc_ids)
        payload["meta"][0]["asset_id"] = asset_id
        workbook_path = self._workbook_path(asset)

        wb = Workbook()
        default_sheet = wb.active
        wb.remove(default_sheet)
        self._write_sheet(wb, "meta", payload["meta"])
        self._write_sheet(wb, "documents", payload["documents"])
        self._write_sheet(wb, "diseases", payload["diseases"])
        self._write_sheet(wb, "sections", payload["sections"])
        self._write_sheet(wb, "chunks", payload["chunks"])
        self._write_sheet(wb, "assertions", payload["assertions"])
        self._write_sheet(wb, "sign_mentions", payload["sign_mentions"])
        self._write_sheet(wb, "service_mentions", payload["service_mentions"])
        self._write_sheet(wb, "observation_mentions", payload["observation_mentions"])
        self._write_sheet(wb, "manual_labels", payload["manual_labels"])
        self._write_sheet(wb, "summaries", payload["summaries"])
        wb.save(workbook_path)

        relative_path = workbook_path.relative_to(BASE_DIR).as_posix()
        return {
            "asset_id": asset_id,
            "workbook_path": str(workbook_path),
            "relative_path": relative_path,
            "download_url": f"/pdfs/{relative_path}",
            "sheet_counts": {key: len(value) for key, value in payload.items() if isinstance(value, list)},
            "exported_at": datetime.now().isoformat(timespec="seconds"),
        }

    def import_asset_workbook(self, asset_id: str, workbook_path: str | Path) -> dict[str, Any]:
        asset = self.knowledge_registry_store.get_asset(asset_id)
        if asset is None:
            raise KeyError(f"Asset not found: {asset_id}")

        workbook_path = Path(workbook_path)
        if not workbook_path.exists():
            raise FileNotFoundError(f"Workbook not found: {workbook_path}")

        wb = load_workbook(workbook_path)
        rows = {name: self._read_sheet_rows(wb[name]) for name in wb.sheetnames}
        meta_rows = rows.get("meta") or []
        if meta_rows:
            meta_asset_id = str(meta_rows[0].get("asset_id") or "").strip()
            if meta_asset_id and meta_asset_id != asset_id:
                raise ValueError(f"Workbook asset_id mismatch: {meta_asset_id} != {asset_id}")

        namespace = self._extract_namespace(rows)
        with self.ontology_store.driver.session() as session:
            summary = {
                "documents": self._sync_documents(session, namespace, rows.get("documents") or []),
                "diseases": self._sync_diseases(session, namespace, rows.get("diseases") or []),
                "sections": self._sync_sections(session, namespace, rows.get("sections") or []),
                "chunks": self._sync_chunks(session, namespace, rows.get("chunks") or []),
                "assertions": self._sync_assertions(session, namespace, rows.get("assertions") or []),
                "sign_mentions": self._sync_sign_mentions(session, namespace, rows.get("sign_mentions") or []),
                "service_mentions": self._sync_service_mentions(session, namespace, rows.get("service_mentions") or []),
                "observation_mentions": self._sync_observation_mentions(session, namespace, rows.get("observation_mentions") or []),
                "summaries": self._sync_summaries(session, namespace, rows.get("summaries") or []),
            }
        manual_rows = rows.get("manual_labels")
        if manual_rows is not None and self.knowledge_text_bridge is not None:
            manual_view = self.knowledge_text_bridge.replace_manual_labels(asset_id, manual_rows)
            summary["manual_labels"] = manual_view.get("manual_label_summary") or {"count": len(manual_rows or [])}
        elif manual_rows is not None:
            normalized_labels = self.knowledge_registry_store.replace_manual_labels(asset_id, manual_rows)
            summary["manual_labels"] = {"count": len(normalized_labels)}

        graph_trace = self.knowledge_registry_store.graph_trace(asset_id)
        return {
            "asset_id": asset_id,
            "namespace": namespace,
            "workbook_path": str(workbook_path),
            "applied": summary,
            "graph_trace": graph_trace,
            "synced_at": datetime.now().isoformat(timespec="seconds"),
        }

    def _collect_asset_payload(self, asset_id: str, doc_ids: list[str]) -> dict[str, list[dict[str, Any]]]:
        with self.ontology_store.driver.session() as session:
            documents = [
                dict(record)
                for record in session.run(
                    """
                    MATCH (doc:RawDocument)
                    WHERE doc.doc_id IN $doc_ids
                    RETURN doc.doc_id AS doc_id,
                           doc.namespace AS namespace,
                           doc.title AS title,
                           doc.file_path AS file_path,
                           doc.source_type AS source_type,
                           doc.doc_type AS doc_type,
                           doc.page_count AS page_count,
                           '' AS __op
                    ORDER BY doc.title, doc.doc_id
                    """,
                    doc_ids=doc_ids,
                )
            ]
            diseases = [
                dict(record)
                for record in session.run(
                    """
                    MATCH (doc:RawDocument)
                    WHERE doc.doc_id IN $doc_ids
                    MATCH (doc)<-[:FROM_DOCUMENT]-(chunk:RawChunk)-[:CHUNK_ABOUT_DISEASE]->(d:DiseaseEntity)
                    RETURN DISTINCT d.disease_id AS disease_id,
                           d.namespace AS namespace,
                           d.disease_name AS disease_name,
                           '' AS __op
                    ORDER BY d.disease_name, d.disease_id
                    """,
                    doc_ids=doc_ids,
                )
            ]
            sections = [
                dict(record)
                for record in session.run(
                    """
                    MATCH (doc:RawDocument)
                    WHERE doc.doc_id IN $doc_ids
                    MATCH (doc)<-[:FROM_DOCUMENT]-(chunk:RawChunk)<-[:SECTION_HAS_CHUNK]-(sec:ProtocolSection)
                    OPTIONAL MATCH (book:ProtocolBook)-[:BOOK_HAS_SECTION]->(sec)
                    RETURN DISTINCT sec.section_id AS section_id,
                           sec.namespace AS namespace,
                           sec.section_title AS section_title,
                           sec.section_type AS section_type,
                           sec.disease_id AS disease_id,
                           sec.page_start AS page_start,
                           sec.page_end AS page_end,
                           book.book_id AS book_id,
                           '' AS __op
                    ORDER BY sec.section_title, sec.section_id
                    """,
                    doc_ids=doc_ids,
                )
            ]
            chunks = [
                dict(record)
                for record in session.run(
                    """
                    MATCH (doc:RawDocument)
                    WHERE doc.doc_id IN $doc_ids
                    MATCH (doc)<-[:FROM_DOCUMENT]-(chunk:RawChunk)
                    OPTIONAL MATCH (sec:ProtocolSection)-[:SECTION_HAS_CHUNK]->(chunk)
                    RETURN chunk.chunk_id AS chunk_id,
                           chunk.namespace AS namespace,
                           chunk.disease_id AS disease_id,
                           chunk.section_type AS section_type,
                           chunk.section_title AS section_title,
                           chunk.parent_section_path AS parent_section_path,
                           coalesce(chunk.page_numbers, []) AS page_numbers,
                           chunk.body_preview AS body_preview,
                           chunk.body_text AS body_text,
                           doc.doc_id AS doc_id,
                           sec.section_id AS section_id,
                           '' AS __op
                    ORDER BY chunk.section_title, chunk.chunk_id
                    """,
                    doc_ids=doc_ids,
                )
            ]
            assertions = [
                dict(record)
                for record in session.run(
                    """
                    MATCH (doc:RawDocument)
                    WHERE doc.doc_id IN $doc_ids
                    MATCH (doc)<-[:FROM_DOCUMENT]-(chunk:RawChunk)-[:CONTAINS_ASSERTION]->(a:ProtocolAssertion)
                    OPTIONAL MATCH (sec:ProtocolSection)-[:CONTAINS_ASSERTION]->(a)
                    OPTIONAL MATCH (a)-[:ASSERTION_ABOUT_DISEASE]->(d:DiseaseEntity)
                    OPTIONAL MATCH (a)-[:ASSERTION_REQUIRES_SIGN]->(sc)
                    WITH doc, chunk, a, sec, d, collect(DISTINCT coalesce(sc.sign_id, sc.claim_sign_id, sc.canonical_label, sc.text)) AS sign_refs
                    OPTIONAL MATCH (a)-[:ASSERTION_INDICATES_SERVICE]->(svc)
                    WITH doc, chunk, a, sec, d, sign_refs, collect(DISTINCT svc.service_code) AS service_refs
                    OPTIONAL MATCH (a)-[:ASSERTION_CONTRAINDICATES]->(csvc)
                    RETURN a.assertion_id AS assertion_id,
                           a.namespace AS namespace,
                           d.disease_id AS disease_id,
                           a.assertion_type AS assertion_type,
                           a.assertion_text AS assertion_text,
                           a.condition_text AS condition_text,
                           a.action_text AS action_text,
                           a.status AS status,
                           a.evidence_level AS evidence_level,
                           a.source_chunk_id AS source_chunk_id,
                           a.source_page AS source_page,
                           sec.section_id AS section_id,
                           sign_refs AS sign_refs,
                           service_refs AS service_refs,
                           collect(DISTINCT csvc.service_code) AS contraindicated_service_refs,
                           '' AS __op
                    ORDER BY a.assertion_id
                    """,
                    doc_ids=doc_ids,
                )
            ]
            sign_mentions = [
                dict(record)
                for record in session.run(
                    """
                    MATCH (doc:RawDocument)
                    WHERE doc.doc_id IN $doc_ids
                    MATCH (doc)<-[:FROM_DOCUMENT]-(chunk:RawChunk)-[:MENTIONS_SIGN]->(m:RawSignMention)
                    OPTIONAL MATCH (m)-[:MAPS_TO_SIGN]->(sc)
                    RETURN m.mention_id AS mention_id,
                           m.namespace AS namespace,
                           m.mention_text AS mention_text,
                           m.context_text AS context_text,
                           m.modifier_raw AS modifier_raw,
                           m.extraction_confidence AS extraction_confidence,
                           m.mapping_status AS mapping_status,
                           m.source_chunk_id AS source_chunk_id,
                           m.source_page AS source_page,
                           coalesce(sc.sign_id, sc.claim_sign_id, sc.canonical_label, sc.text) AS concept_ref,
                           coalesce(m.manual_override, false) AS manual_override,
                           coalesce(m.manual_note, '') AS manual_note,
                           coalesce(m.manual_label_id, '') AS manual_label_id,
                           coalesce(m.chunk_match_mode, '') AS chunk_match_mode,
                           '' AS __op
                    ORDER BY m.mention_id
                    """,
                    doc_ids=doc_ids,
                )
            ]
            service_mentions = [
                dict(record)
                for record in session.run(
                    """
                    MATCH (doc:RawDocument)
                    WHERE doc.doc_id IN $doc_ids
                    MATCH (doc)<-[:FROM_DOCUMENT]-(chunk:RawChunk)-[:MENTIONS_SERVICE]->(m:RawServiceMention)
                    OPTIONAL MATCH (m)-[:MAPS_TO_SERVICE]->(svc)
                    RETURN m.mention_id AS mention_id,
                           m.namespace AS namespace,
                           m.mention_text AS mention_text,
                           m.context_text AS context_text,
                           m.medical_role AS medical_role,
                           m.condition_to_apply AS condition_to_apply,
                           m.extraction_confidence AS extraction_confidence,
                           m.mapping_status AS mapping_status,
                           m.source_chunk_id AS source_chunk_id,
                           m.source_page AS source_page,
                           svc.service_code AS service_code,
                           coalesce(m.manual_override, false) AS manual_override,
                           coalesce(m.manual_note, '') AS manual_note,
                           coalesce(m.manual_label_id, '') AS manual_label_id,
                           coalesce(m.chunk_match_mode, '') AS chunk_match_mode,
                           '' AS __op
                    ORDER BY m.mention_id
                    """,
                    doc_ids=doc_ids,
                )
            ]
            observation_mentions = [
                dict(record)
                for record in session.run(
                    """
                    MATCH (doc:RawDocument)
                    WHERE doc.doc_id IN $doc_ids
                    MATCH (doc)<-[:FROM_DOCUMENT]-(chunk:RawChunk)-[:MENTIONS_OBSERVATION]->(m:RawObservationMention)
                    RETURN m.mention_id AS mention_id,
                           m.namespace AS namespace,
                           m.mention_text AS mention_text,
                           m.context_text AS context_text,
                           m.result_semantics AS result_semantics,
                           m.extraction_confidence AS extraction_confidence,
                           m.mapping_status AS mapping_status,
                           m.source_chunk_id AS source_chunk_id,
                           m.source_page AS source_page,
                           coalesce(m.manual_override, false) AS manual_override,
                           coalesce(m.manual_note, '') AS manual_note,
                           coalesce(m.manual_label_id, '') AS manual_label_id,
                           coalesce(m.chunk_match_mode, '') AS chunk_match_mode,
                           '' AS __op
                    ORDER BY m.mention_id
                    """,
                    doc_ids=doc_ids,
                )
            ]
            summaries = [
                dict(record)
                for record in session.run(
                    """
                    MATCH (doc:RawDocument)
                    WHERE doc.doc_id IN $doc_ids
                    MATCH (doc)<-[:FROM_DOCUMENT]-(chunk:RawChunk)-[:CHUNK_ABOUT_DISEASE]->(d:DiseaseEntity)
                    MATCH (s:ProtocolDiseaseSummary)-[:SUMMARIZES]->(d)
                    RETURN DISTINCT s.summary_id AS summary_id,
                           s.namespace AS namespace,
                           d.disease_id AS disease_id,
                           s.summary_text AS summary_text,
                           coalesce(s.key_signs, []) AS key_signs,
                           coalesce(s.key_services, []) AS key_services,
                           coalesce(s.key_drugs, []) AS key_drugs,
                           coalesce(s.differential_diseases, []) AS differential_diseases,
                           '' AS __op
                    ORDER BY s.summary_id
                    """,
                    doc_ids=doc_ids,
                )
            ]

        meta = [
            {
                "asset_id": "",
                "namespace": documents[0].get("namespace") if documents else "",
                "doc_ids_json": json.dumps(doc_ids, ensure_ascii=False),
                "exported_at": datetime.now().isoformat(timespec="seconds"),
                "note": "Sheet co cot __op: blank=giu/update, create=them moi, delete=xoa. Co the sua nhan o mention sheets hoac manual_labels.",
            }
        ]
        return {
            "meta": meta,
            "documents": self._normalize_rows(documents),
            "diseases": self._normalize_rows(diseases),
            "sections": self._normalize_rows(sections),
            "chunks": self._normalize_rows(chunks),
            "assertions": self._normalize_rows(assertions),
            "sign_mentions": self._normalize_rows(sign_mentions),
            "service_mentions": self._normalize_rows(service_mentions),
            "observation_mentions": self._normalize_rows(observation_mentions),
            "manual_labels": self._normalize_rows(self._manual_label_rows(asset_id)),
            "summaries": self._normalize_rows(summaries),
        }

    def _manual_label_rows(self, asset_id: str) -> list[dict[str, Any]]:
        rows: list[dict[str, Any]] = []
        for item in self.knowledge_registry_store.list_manual_labels(asset_id):
            rows.append(
                {
                    "label_id": item.get("label_id") or "",
                    "mention_id": item.get("mention_id") or "",
                    "kind": item.get("kind") or "",
                    "text": item.get("text") or "",
                    "concept_ref": item.get("concept_ref") or "",
                    "concept_label": item.get("concept_label") or "",
                    "note": item.get("note") or "",
                    "source_chunk_id": item.get("source_chunk_id") or "",
                    "source_page": item.get("source_page"),
                    "start_offset": item.get("start_offset"),
                    "end_offset": item.get("end_offset"),
                    "manual_status": item.get("manual_status") or "active",
                    "__op": "",
                }
            )
        return rows

    def _normalize_rows(self, rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
        normalized = []
        for row in rows:
            normalized.append({key: self._serialize_cell(value) for key, value in row.items()})
        return normalized

    def _serialize_cell(self, value: Any) -> Any:
        if isinstance(value, list):
            return json.dumps(value, ensure_ascii=False)
        return value

    def _write_sheet(self, wb: Workbook, title: str, rows: list[dict[str, Any]]) -> None:
        ws = wb.create_sheet(title=title[:31])
        if not rows:
            headers = self._empty_sheet_headers(title)
            ws.append(headers or ["empty"])
            for cell in ws[1]:
                cell.font = Font(bold=True)
            ws.freeze_panes = "A2"
            return
        headers = list(rows[0].keys())
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)
        for row in rows:
            ws.append([row.get(header, "") for header in headers])
        end_col = self._column_letter(len(headers))
        end_row = max(len(rows) + 1, 2)
        tab = Table(displayName=f"tbl_{title[:20]}", ref=f"A1:{end_col}{end_row}")
        tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False)
        ws.add_table(tab)
        ws.freeze_panes = "A2"

    def _empty_sheet_headers(self, title: str) -> list[str]:
        schema = {
            "manual_labels": [
                "label_id",
                "mention_id",
                "kind",
                "text",
                "concept_ref",
                "concept_label",
                "note",
                "source_chunk_id",
                "source_page",
                "start_offset",
                "end_offset",
                "manual_status",
                "__op",
            ],
        }
        return schema.get(title, [])

    def _read_sheet_rows(self, ws) -> list[dict[str, Any]]:
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(item or "").strip() for item in rows[0]]
        result: list[dict[str, Any]] = []
        for values in rows[1:]:
            if not any(value not in (None, "") for value in values):
                continue
            row = {}
            for index, header in enumerate(headers):
                if not header:
                    continue
                row[header] = values[index] if index < len(values) else None
            result.append(row)
        return result

    def _extract_namespace(self, rows: dict[str, list[dict[str, Any]]]) -> str:
        for sheet_name in ("documents", "diseases", "sections", "chunks"):
            for row in rows.get(sheet_name) or []:
                namespace = str(row.get("namespace") or "").strip()
                if namespace:
                    return namespace
        return "ontology_v2"

    def _sync_documents(self, session, namespace: str, rows: list[dict[str, Any]]) -> dict[str, int]:
        updated = 0
        for row in rows:
            doc_id = str(row.get("doc_id") or "").strip()
            if not doc_id:
                continue
            session.run(
                """
                MATCH (doc:RawDocument {doc_id:$doc_id})
                SET doc.title = $title,
                    doc.file_path = $file_path,
                    doc.source_type = $source_type,
                    doc.doc_type = $doc_type,
                    doc.page_count = $page_count
                """,
                doc_id=doc_id,
                title=row.get("title") or "",
                file_path=row.get("file_path") or "",
                source_type=row.get("source_type") or "",
                doc_type=row.get("doc_type") or "",
                page_count=self._int_or_none(row.get("page_count")),
            )
            updated += 1
        return {"updated": updated}

    def _sync_diseases(self, session, namespace: str, rows: list[dict[str, Any]]) -> dict[str, int]:
        updated = 0
        for row in rows:
            disease_id = str(row.get("disease_id") or "").strip()
            if not disease_id:
                continue
            session.run(
                """
                MATCH (d:DiseaseEntity {disease_id:$disease_id})
                SET d.disease_name = $disease_name,
                    d.namespace = $namespace
                """,
                disease_id=disease_id,
                disease_name=row.get("disease_name") or disease_id,
                namespace=row.get("namespace") or namespace,
            )
            updated += 1
        return {"updated": updated}

    def _sync_sections(self, session, namespace: str, rows: list[dict[str, Any]]) -> dict[str, int]:
        created = updated = deleted = 0
        for row in rows:
            section_id = str(row.get("section_id") or "").strip()
            if not section_id:
                continue
            op = self._op(row)
            if op == "delete":
                session.run("MATCH (s:ProtocolSection {section_id:$section_id}) DETACH DELETE s", section_id=section_id)
                deleted += 1
                continue
            exists = self._node_exists(session, "ProtocolSection", "section_id", section_id)
            session.run(
                """
                MERGE (s:ProtocolSection {section_id:$section_id})
                SET s.namespace = $namespace,
                    s.section_title = $section_title,
                    s.section_type = $section_type,
                    s.disease_id = $disease_id,
                    s.page_start = $page_start,
                    s.page_end = $page_end
                """,
                section_id=section_id,
                namespace=row.get("namespace") or namespace,
                section_title=row.get("section_title") or "",
                section_type=row.get("section_type") or "",
                disease_id=row.get("disease_id") or "",
                page_start=self._int_or_none(row.get("page_start")),
                page_end=self._int_or_none(row.get("page_end")),
            )
            if row.get("book_id"):
                session.run(
                    """
                    MATCH (b:ProtocolBook {book_id:$book_id})
                    MATCH (s:ProtocolSection {section_id:$section_id})
                    MERGE (b)-[:BOOK_HAS_SECTION]->(s)
                    """,
                    book_id=row.get("book_id"),
                    section_id=section_id,
                )
            if row.get("disease_id"):
                session.run(
                    """
                    MATCH (s:ProtocolSection {section_id:$section_id})
                    MATCH (d:DiseaseEntity {disease_id:$disease_id})
                    MERGE (s)-[:SECTION_COVERS_DISEASE]->(d)
                    """,
                    section_id=section_id,
                    disease_id=row.get("disease_id"),
                )
            created += 0 if exists else 1
            updated += 1 if exists else 0
        return {"created": created, "updated": updated, "deleted": deleted}

    def _sync_chunks(self, session, namespace: str, rows: list[dict[str, Any]]) -> dict[str, int]:
        created = updated = deleted = 0
        for row in rows:
            chunk_id = str(row.get("chunk_id") or "").strip()
            if not chunk_id:
                continue
            op = self._op(row)
            if op == "delete":
                session.run(
                    """
                    MATCH (c:RawChunk {chunk_id:$chunk_id})
                    OPTIONAL MATCH (c)-[:MENTIONS_SIGN]->(sm:RawSignMention)
                    OPTIONAL MATCH (c)-[:MENTIONS_SERVICE]->(svm:RawServiceMention)
                    OPTIONAL MATCH (c)-[:MENTIONS_OBSERVATION]->(om:RawObservationMention)
                    OPTIONAL MATCH (c)-[:CONTAINS_ASSERTION]->(a:ProtocolAssertion)
                    DETACH DELETE c, sm, svm, om, a
                    """,
                    chunk_id=chunk_id,
                )
                deleted += 1
                continue
            exists = self._node_exists(session, "RawChunk", "chunk_id", chunk_id)
            page_numbers = self._json_list(row.get("page_numbers"))
            session.run(
                """
                MERGE (c:RawChunk {chunk_id:$chunk_id})
                SET c.namespace = $namespace,
                    c.disease_id = $disease_id,
                    c.section_type = $section_type,
                    c.section_title = $section_title,
                    c.parent_section_path = $parent_section_path,
                    c.page_numbers = $page_numbers,
                    c.body_preview = $body_preview,
                    c.body_text = $body_text
                """,
                chunk_id=chunk_id,
                namespace=row.get("namespace") or namespace,
                disease_id=row.get("disease_id") or "",
                section_type=row.get("section_type") or "",
                section_title=row.get("section_title") or "",
                parent_section_path=row.get("parent_section_path") or "",
                page_numbers=page_numbers,
                body_preview=row.get("body_preview") or "",
                body_text=row.get("body_text") or "",
            )
            if row.get("doc_id"):
                session.run(
                    """
                    MATCH (c:RawChunk {chunk_id:$chunk_id})
                    MATCH (doc:RawDocument {doc_id:$doc_id})
                    MERGE (c)-[:FROM_DOCUMENT]->(doc)
                    """,
                    chunk_id=chunk_id,
                    doc_id=row.get("doc_id"),
                )
            if row.get("disease_id"):
                session.run(
                    """
                    MATCH (c:RawChunk {chunk_id:$chunk_id})
                    MATCH (d:DiseaseEntity {disease_id:$disease_id})
                    MERGE (c)-[:CHUNK_ABOUT_DISEASE]->(d)
                    """,
                    chunk_id=chunk_id,
                    disease_id=row.get("disease_id"),
                )
            if row.get("section_id"):
                session.run(
                    """
                    MATCH (s:ProtocolSection {section_id:$section_id})
                    MATCH (c:RawChunk {chunk_id:$chunk_id})
                    MERGE (s)-[:SECTION_HAS_CHUNK]->(c)
                    """,
                    section_id=row.get("section_id"),
                    chunk_id=chunk_id,
                )
            created += 0 if exists else 1
            updated += 1 if exists else 0
        return {"created": created, "updated": updated, "deleted": deleted}

    def _sync_assertions(self, session, namespace: str, rows: list[dict[str, Any]]) -> dict[str, int]:
        created = updated = deleted = 0
        for row in rows:
            assertion_id = str(row.get("assertion_id") or "").strip()
            if not assertion_id:
                continue
            op = self._op(row)
            if op == "delete":
                session.run("MATCH (a:ProtocolAssertion {assertion_id:$assertion_id}) DETACH DELETE a", assertion_id=assertion_id)
                deleted += 1
                continue
            exists = self._node_exists(session, "ProtocolAssertion", "assertion_id", assertion_id)
            session.run(
                """
                MERGE (a:ProtocolAssertion {assertion_id:$assertion_id})
                SET a.namespace = $namespace,
                    a.assertion_type = $assertion_type,
                    a.assertion_text = $assertion_text,
                    a.condition_text = $condition_text,
                    a.action_text = $action_text,
                    a.status = $status,
                    a.evidence_level = $evidence_level,
                    a.source_chunk_id = $source_chunk_id,
                    a.source_page = $source_page
                """,
                assertion_id=assertion_id,
                namespace=row.get("namespace") or namespace,
                assertion_type=row.get("assertion_type") or "",
                assertion_text=row.get("assertion_text") or "",
                condition_text=row.get("condition_text") or "",
                action_text=row.get("action_text") or "",
                status=row.get("status") or "",
                evidence_level=row.get("evidence_level") or "",
                source_chunk_id=row.get("source_chunk_id") or "",
                source_page=self._int_or_none(row.get("source_page")),
            )
            if row.get("source_chunk_id"):
                session.run(
                    """
                    MATCH (c:RawChunk {chunk_id:$chunk_id})
                    MATCH (a:ProtocolAssertion {assertion_id:$assertion_id})
                    MERGE (c)-[:CONTAINS_ASSERTION]->(a)
                    """,
                    chunk_id=row.get("source_chunk_id"),
                    assertion_id=assertion_id,
                )
            if row.get("disease_id"):
                session.run(
                    """
                    MATCH (a:ProtocolAssertion {assertion_id:$assertion_id})
                    MATCH (d:DiseaseEntity {disease_id:$disease_id})
                    MERGE (a)-[:ASSERTION_ABOUT_DISEASE]->(d)
                    """,
                    assertion_id=assertion_id,
                    disease_id=row.get("disease_id"),
                )
            if row.get("section_id"):
                session.run(
                    """
                    MATCH (s:ProtocolSection {section_id:$section_id})
                    MATCH (a:ProtocolAssertion {assertion_id:$assertion_id})
                    MERGE (s)-[:CONTAINS_ASSERTION]->(a)
                    """,
                    section_id=row.get("section_id"),
                    assertion_id=assertion_id,
                )
            session.run(
                """
                MATCH (a:ProtocolAssertion {assertion_id:$assertion_id})-[r:ASSERTION_REQUIRES_SIGN]->()
                DELETE r
                """,
                assertion_id=assertion_id,
            )
            for sign_ref in self._json_list(row.get("sign_refs")):
                session.run(
                    """
                    MATCH (a:ProtocolAssertion {assertion_id:$assertion_id})
                    MATCH (s)
                    WHERE (s:SignConcept OR s:ClaimSignConcept OR s:CISign)
                      AND coalesce(s.sign_id, s.claim_sign_id, s.canonical_label, s.text) = $sign_ref
                    MERGE (a)-[:ASSERTION_REQUIRES_SIGN]->(s)
                    """,
                    assertion_id=assertion_id,
                    sign_ref=sign_ref,
                )
            session.run(
                """
                MATCH (a:ProtocolAssertion {assertion_id:$assertion_id})-[r:ASSERTION_INDICATES_SERVICE|ASSERTION_CONTRAINDICATES]->()
                DELETE r
                """,
                assertion_id=assertion_id,
            )
            for service_code in self._json_list(row.get("service_refs")):
                session.run(
                    """
                    MATCH (a:ProtocolAssertion {assertion_id:$assertion_id})
                    MATCH (svc)
                    WHERE (svc:ProtocolService OR svc:CIService)
                      AND svc.service_code = $service_code
                    MERGE (a)-[:ASSERTION_INDICATES_SERVICE]->(svc)
                    """,
                    assertion_id=assertion_id,
                    service_code=service_code,
                )
            for service_code in self._json_list(row.get("contraindicated_service_refs")):
                session.run(
                    """
                    MATCH (a:ProtocolAssertion {assertion_id:$assertion_id})
                    MATCH (svc)
                    WHERE (svc:ProtocolService OR svc:CIService)
                      AND svc.service_code = $service_code
                    MERGE (a)-[:ASSERTION_CONTRAINDICATES]->(svc)
                    """,
                    assertion_id=assertion_id,
                    service_code=service_code,
                )
            created += 0 if exists else 1
            updated += 1 if exists else 0
        return {"created": created, "updated": updated, "deleted": deleted}

    def _sync_sign_mentions(self, session, namespace: str, rows: list[dict[str, Any]]) -> dict[str, int]:
        return self._sync_mentions(
            session=session,
            namespace=namespace,
            rows=rows,
            label="RawSignMention",
            chunk_rel="MENTIONS_SIGN",
            id_key="mention_id",
            property_map={
                "mention_text": "mention_text",
                "context_text": "context_text",
                "modifier_raw": "modifier_raw",
                "extraction_confidence": "extraction_confidence",
                "mapping_status": "mapping_status",
                "source_chunk_id": "source_chunk_id",
                "source_page": "source_page",
                "manual_override": "manual_override",
                "manual_note": "manual_note",
                "manual_label_id": "manual_label_id",
                "chunk_match_mode": "chunk_match_mode",
            },
            concept_field="concept_ref",
            concept_query="""
                MATCH (m:RawSignMention {mention_id:$mention_id})-[r:MAPS_TO_SIGN]->()
                DELETE r
            """,
            concept_merge="""
                MATCH (m:RawSignMention {mention_id:$mention_id})
                MATCH (s)
                WHERE (s:SignConcept OR s:ClaimSignConcept OR s:CISign)
                  AND coalesce(s.sign_id, s.claim_sign_id, s.canonical_label, s.text) = $concept_ref
                MERGE (m)-[:MAPS_TO_SIGN]->(s)
            """,
        )

    def _sync_service_mentions(self, session, namespace: str, rows: list[dict[str, Any]]) -> dict[str, int]:
        return self._sync_mentions(
            session=session,
            namespace=namespace,
            rows=rows,
            label="RawServiceMention",
            chunk_rel="MENTIONS_SERVICE",
            id_key="mention_id",
            property_map={
                "mention_text": "mention_text",
                "context_text": "context_text",
                "medical_role": "medical_role",
                "condition_to_apply": "condition_to_apply",
                "extraction_confidence": "extraction_confidence",
                "mapping_status": "mapping_status",
                "source_chunk_id": "source_chunk_id",
                "source_page": "source_page",
                "manual_override": "manual_override",
                "manual_note": "manual_note",
                "manual_label_id": "manual_label_id",
                "chunk_match_mode": "chunk_match_mode",
            },
            concept_field="service_code",
            concept_query="""
                MATCH (m:RawServiceMention {mention_id:$mention_id})-[r:MAPS_TO_SERVICE]->()
                DELETE r
            """,
            concept_merge="""
                MATCH (m:RawServiceMention {mention_id:$mention_id})
                MATCH (svc)
                WHERE (svc:ProtocolService OR svc:CIService)
                  AND svc.service_code = $concept_ref
                MERGE (m)-[:MAPS_TO_SERVICE]->(svc)
            """,
        )

    def _sync_observation_mentions(self, session, namespace: str, rows: list[dict[str, Any]]) -> dict[str, int]:
        return self._sync_mentions(
            session=session,
            namespace=namespace,
            rows=rows,
            label="RawObservationMention",
            chunk_rel="MENTIONS_OBSERVATION",
            id_key="mention_id",
            property_map={
                "mention_text": "mention_text",
                "context_text": "context_text",
                "result_semantics": "result_semantics",
                "extraction_confidence": "extraction_confidence",
                "mapping_status": "mapping_status",
                "source_chunk_id": "source_chunk_id",
                "source_page": "source_page",
                "manual_override": "manual_override",
                "manual_note": "manual_note",
                "manual_label_id": "manual_label_id",
                "chunk_match_mode": "chunk_match_mode",
            },
        )

    def _sync_mentions(
        self,
        *,
        session,
        namespace: str,
        rows: list[dict[str, Any]],
        label: str,
        chunk_rel: str,
        id_key: str,
        property_map: dict[str, str],
        concept_field: str | None = None,
        concept_query: str | None = None,
        concept_merge: str | None = None,
    ) -> dict[str, int]:
        created = updated = deleted = 0
        for row in rows:
            mention_id = str(row.get(id_key) or "").strip()
            if not mention_id:
                continue
            op = self._op(row)
            if op == "delete":
                session.run(f"MATCH (m:{label} {{{id_key}:$mention_id}}) DETACH DELETE m", mention_id=mention_id)
                deleted += 1
                continue
            exists = self._node_exists(session, label, id_key, mention_id)
            assignments = ["m.namespace = $namespace"]
            params = {"mention_id": mention_id, "namespace": row.get("namespace") or namespace}
            for prop, source in property_map.items():
                assignments.append(f"m.{prop} = ${prop}")
                params[prop] = row.get(source)
            cypher = f"MERGE (m:{label} {{{id_key}:$mention_id}}) SET " + ", ".join(assignments)
            session.run(cypher, **params)
            if row.get("source_chunk_id"):
                session.run(
                    f"""
                    MATCH (c:RawChunk {{chunk_id:$chunk_id}})
                    MATCH (m:{label} {{{id_key}:$mention_id}})
                    MERGE (c)-[:{chunk_rel}]->(m)
                    """,
                    chunk_id=row.get("source_chunk_id"),
                    mention_id=mention_id,
                )
            if concept_query:
                session.run(concept_query, mention_id=mention_id)
            if concept_merge and concept_field and row.get(concept_field):
                session.run(concept_merge, mention_id=mention_id, concept_ref=row.get(concept_field))
            created += 0 if exists else 1
            updated += 1 if exists else 0
        return {"created": created, "updated": updated, "deleted": deleted}

    def _sync_summaries(self, session, namespace: str, rows: list[dict[str, Any]]) -> dict[str, int]:
        created = updated = deleted = 0
        for row in rows:
            summary_id = str(row.get("summary_id") or "").strip()
            if not summary_id:
                continue
            op = self._op(row)
            if op == "delete":
                session.run("MATCH (s:ProtocolDiseaseSummary {summary_id:$summary_id}) DETACH DELETE s", summary_id=summary_id)
                deleted += 1
                continue
            exists = self._node_exists(session, "ProtocolDiseaseSummary", "summary_id", summary_id)
            session.run(
                """
                MERGE (s:ProtocolDiseaseSummary {summary_id:$summary_id})
                SET s.namespace = $namespace,
                    s.summary_text = $summary_text,
                    s.key_signs = $key_signs,
                    s.key_services = $key_services,
                    s.key_drugs = $key_drugs,
                    s.differential_diseases = $differential_diseases
                """,
                summary_id=summary_id,
                namespace=row.get("namespace") or namespace,
                summary_text=row.get("summary_text") or "",
                key_signs=self._json_list(row.get("key_signs")),
                key_services=self._json_list(row.get("key_services")),
                key_drugs=self._json_list(row.get("key_drugs")),
                differential_diseases=self._json_list(row.get("differential_diseases")),
            )
            if row.get("disease_id"):
                session.run(
                    """
                    MATCH (s:ProtocolDiseaseSummary {summary_id:$summary_id})
                    MATCH (d:DiseaseEntity {disease_id:$disease_id})
                    MERGE (s)-[:SUMMARIZES]->(d)
                    """,
                    summary_id=summary_id,
                    disease_id=row.get("disease_id"),
                )
            created += 0 if exists else 1
            updated += 1 if exists else 0
        return {"created": created, "updated": updated, "deleted": deleted}

    def _node_exists(self, session, label: str, prop: str, value: str) -> bool:
        row = session.run(
            f"MATCH (n:{label} {{{prop}:$value}}) RETURN count(n) AS c",
            value=value,
        ).single()
        return bool(row and row["c"])

    def _op(self, row: dict[str, Any]) -> str:
        return str(row.get("__op") or "").strip().lower()

    def _json_list(self, value: Any) -> list[Any]:
        if value in (None, ""):
            return []
        if isinstance(value, list):
            return value
        if isinstance(value, str):
            raw = value.strip()
            if not raw:
                return []
            try:
                parsed = json.loads(raw)
                if isinstance(parsed, list):
                    return parsed
            except json.JSONDecodeError:
                return [item.strip() for item in raw.split("|") if item.strip()]
        return [value]

    def _int_or_none(self, value: Any) -> int | None:
        if value in (None, ""):
            return None
        try:
            return int(value)
        except (TypeError, ValueError):
            return None

    def _workbook_path(self, asset: dict[str, Any]) -> Path:
        title = str(asset.get("title") or Path(asset.get("source_path") or "asset").stem)
        safe_title = title.replace(" ", "_").replace("/", "_").replace("\\", "_")
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        return KNOWLEDGE_EXCEL_VIEWS_DIR / f"{safe_title}_{stamp}.xlsx"

    def _column_letter(self, index: int) -> str:
        letters = ""
        while index:
            index, rem = divmod(index - 1, 26)
            letters = chr(65 + rem) + letters
        return letters or "A"
