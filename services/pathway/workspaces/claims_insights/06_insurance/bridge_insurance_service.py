"""Bridge Insurance ↔ Service — connect 3 domains in Neo4j.

Three bridges:
  A. Plan Resolution:  BaoCaoHoSo real plan names → benefit pack plans
  B. Category→Benefit: CIService category_code → Benefit entries
  C. Historical Exclusion: CIService → Exclusion patterns (from claim signals)

Usage:
    set PYTHONIOENCODING=utf-8 && python -X utf8 bridge_insurance_service.py [--clear]
"""

from __future__ import annotations

import argparse
import hashlib
import json
import logging
import re
import unicodedata
from collections import Counter
from pathlib import Path
from typing import Any

import openpyxl
from neo4j import GraphDatabase

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s",
)
logger = logging.getLogger(__name__)

# ── Paths ──────────────────────────────────────────────────────────
_HERE = Path(__file__).resolve().parent
_CLAIMS_ROOT = _HERE.parent  # workspaces/claims_insights
BAOCAOHOSO_PATH = _CLAIMS_ROOT / "01_claims" / "BaoCaoHoSo.xlsx"
CODEBOOK_PATH = _CLAIMS_ROOT / "02_standardize" / "service_codebook.json"
BENEFIT_PACK_PATH = _HERE / "benefit_contract_knowledge_pack.json"
EXCLUSION_SIGNAL_PATH = _HERE / "contract_clause_service_catalog.json"
EXCLUSION_PACK_PATH = _HERE / "exclusion_knowledge_pack.json"
INSURANCE_NAMESPACE = "claims_insights_insurance_v1"

NEO4J_URI = "bolt://localhost:7688"
NEO4J_AUTH = ("neo4j", "password123")


# =====================================================================
# Bridge A: Plan Resolution
# =====================================================================

# Map BaoCaoHoSo plan names (partial match) → benefit pack contract+plan
# The plan names in BaoCaoHoSo are long, e.g.:
#   "Gói D (Từ 15 ngày tuổi tới tròn 65 tuổi)"
#   "Chương trình bảo hiểm cho nhân viên cấp 3 (từ 18 đến 65 tuổi)"
# We normalize to match benefit pack plan names.

PLAN_RESOLUTION_RULES = [
    # FPT-NT plans: "Gói X (...)" → FPT-NT-2025 Gói X
    # Use 2025 as default (latest); 2024 is legacy
    # Match "Gói mở rộng" before "Gói" to avoid partial match
    (r"gói mở rộng",  "FPT-NT-2025", "Mở rộng"),
    (r"gói a\b",      "FPT-NT-2025", "Gói A"),
    (r"gói b\b",      "FPT-NT-2025", "Gói B"),
    (r"gói c\b",      "FPT-NT-2025", "Gói C"),
    (r"gói d\b",      "FPT-NT-2025", "Gói D"),
    # FPT-NV plans: "NV cấp X (...)" → FPT-NV CBNV Cấp X
    (r"nhân viên cấp 1-2|nhân viên cấp 1\b", "FPT-NV", "CBNV Cấp 1-2"),
    (r"nhân viên cấp 3\b",                    "FPT-NV", "CBNV Cấp 3"),
    (r"nhân viên cấp 4\b",                    "FPT-NV", "CBNV Cấp 4"),
]


def resolve_plan_name(raw_plan: str) -> dict[str, str] | None:
    """Resolve a BaoCaoHoSo plan name to benefit pack contract+plan."""
    if not raw_plan:
        return None
    lower = raw_plan.lower()
    for pattern, contract_id, plan_name in PLAN_RESOLUTION_RULES:
        if re.search(pattern, lower, re.IGNORECASE):
            return {"contract_id": contract_id, "plan_name": plan_name}
    return None


# Map insurer full names → short codes used in contract_rules
INSURER_NORMALIZE = {
    "petrolimex": "PJICO",
    "pjico": "PJICO",
    "techcom": "TCGIns",
    "liên hiệp": "UIC",
    "uic": "UIC",
    "bảo hiểm hùng": "BHV",
    "dbv": "DBV",
    "tokio": "TM",
    "kỷ nguyên": "KN",
}


def normalize_insurer(raw: str) -> str:
    """Normalize insurer name to short code."""
    lower = raw.lower()
    for key, code in INSURER_NORMALIZE.items():
        if key in lower:
            return code
    return raw[:20]


# =====================================================================
# Bridge B: CIService Category → Benefit
# =====================================================================

# All subclinical service categories map to a common benefit entry.
# The benefit differentiation is at PLAN level (limits), not category level.
# Dental services are the exception — they have a separate benefit.

CATEGORY_TO_BENEFIT_GROUP = {
    # Subclinical — all under "khám, xét nghiệm, chẩn đoán"
    "LAB-BIO": "Chi phí khám, xét nghiệm, chẩn đoán",
    "LAB-IMM": "Chi phí khám, xét nghiệm, chẩn đoán",
    "LAB-HEM": "Chi phí khám, xét nghiệm, chẩn đoán",
    "LAB-MIC": "Chi phí khám, xét nghiệm, chẩn đoán",
    "LAB-URI": "Chi phí khám, xét nghiệm, chẩn đoán",
    "IMG-XRY": "Chi phí khám, xét nghiệm, chẩn đoán",
    "IMG-USG": "Chi phí khám, xét nghiệm, chẩn đoán",
    "IMG-CTN": "Chi phí khám, xét nghiệm, chẩn đoán",
    "END-ENS": "Chi phí khám, xét nghiệm, chẩn đoán",
    "FUN-DFT": "Chi phí khám, xét nghiệm, chẩn đoán",
    "PAT-PAT": "Chi phí khám, xét nghiệm, chẩn đoán",
    "PRO-THT": "Chi phí khám, xét nghiệm, chẩn đoán",
    "GEN-OTH": "Chi phí khám, xét nghiệm, chẩn đoán",
}

TOKEN_STOPWORDS = {
    "chi", "phi", "dich", "vu", "khong", "thuoc", "quyen", "loi", "hop", "dong",
    "bao", "hiem", "nguoi", "benh", "dieu", "tri", "ngoai", "tru", "noi", "tru",
    "muc", "gioi", "han", "lan", "kham", "theo", "cua", "va", "la", "cac", "mot",
    "cho", "tai", "neu", "voi", "den", "khi", "co", "do", "duoc", "nay", "tu",
    "phan", "ung", "nhung", "tren", "duoi", "can", "hoso", "ho", "so", "te", "y",
    "xet", "nghiem", "chan", "doan", "thuoc", "toa", "nguoi", "nam", "per",
    "visit", "basic", "support", "service", "cost", "outpatient", "benefit",
}


def _text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _ascii_fold(value: Any) -> str:
    text = _text(value).lower()
    text = text.replace("đ", "d").replace("Đ", "d")
    normalized = unicodedata.normalize("NFKD", text)
    return "".join(ch for ch in normalized if not unicodedata.combining(ch))


def _tokenize(value: Any) -> list[str]:
    folded = _ascii_fold(value)
    tokens = re.findall(r"[a-z0-9]+", folded)
    return [
        token for token in tokens
        if len(token) >= 4 and token not in TOKEN_STOPWORDS
    ]


def _unique_ordered(values: list[str]) -> list[str]:
    seen: set[str] = set()
    result: list[str] = []
    for value in values:
        if not value or value in seen:
            continue
        seen.add(value)
        result.append(value)
    return result


def _score_overlap(source_text: str, target_text: str) -> tuple[float, list[str]]:
    source_folded = _ascii_fold(source_text)
    target_folded = _ascii_fold(target_text)
    if not source_folded or not target_folded:
        return 0.0, []

    source_tokens = _unique_ordered(_tokenize(source_text))
    if not source_tokens:
        return 0.0, []

    matched_terms = [token for token in source_tokens if token in target_folded]
    if not matched_terms:
        return 0.0, []

    score = len(matched_terms) / max(len(source_tokens), 1)
    if source_folded in target_folded:
        score = max(score, 0.96)
    elif len(matched_terms) >= 3:
        score = max(score, 0.65)
    elif len(matched_terms) >= 2:
        score = max(score, 0.48)
    return round(min(score, 1.0), 4), matched_terms[:8]


def _classify_exclusion_resolution(code: str, group: str, reason: str) -> str:
    code_norm = _ascii_fold(code)
    text = f"{_ascii_fold(group)} {_ascii_fold(reason)}"

    financial_codes = {"ma01", "ma04", "ma11", "ma12", "ld0001", "ma003"}
    documentation_codes = {"ma07", "ma08", "ma09", "ma10", "ma19"}
    hard_deny_codes = {"ma05", "ma15", "ma21", "ma26", "ma27", "ma28", "ma30", "ma004"}
    soft_review_codes = {"ma06", "ma39"}

    if code_norm in financial_codes:
        return "financial"
    if code_norm in documentation_codes:
        return "documentation"
    if code_norm in hard_deny_codes:
        return "hard_deny"
    if code_norm in soft_review_codes:
        return "soft_review"
    if "dong chi tra" in text or "han muc" in text or "giam tru" in text:
        return "financial"
    if "chung tu" in text or "hoa don" in text:
        return "documentation"
    if "tam soat" in text or "kiem tra" in text:
        return "soft_review"
    if "khong thuoc quyen loi" in text or "loai tru" in text:
        return "hard_deny"
    return "review"


def _is_generic_benefit(entry_label: str, canonical_name: str) -> bool:
    text = _ascii_fold(f"{entry_label} {canonical_name}")
    generic_markers = [
        "dieu tri ngoai tru",
        "dieu tri noi tru",
        "quyen loi co ban",
        "gioi han/lan kham",
        "han muc nguoi/nam",
        "chi phi dieu tri ngoai tru",
        "chi phi dieu tri noi tru",
    ]
    if any(marker in text for marker in generic_markers):
        return True
    if re.match(r"^\d+\.\s", _text(entry_label)):
        return True
    return False


# =====================================================================
# Bridge C: Historical Exclusion → CIService
# =====================================================================
# Data source: contract_clause_service_catalog.json
# Contains: reason → top_services with service_code + row counts
# We create: (:CIService)-[:HISTORICALLY_EXCLUDED {rows, reason, contract}]->(:ExclusionPattern)


# =====================================================================
# Main Engine
# =====================================================================

class InsuranceServiceBridge:
    """Build 3 bridges connecting insurance domain to service domain in Neo4j."""

    def __init__(self):
        self.driver = GraphDatabase.driver(NEO4J_URI, auth=NEO4J_AUTH)
        self.stats = {
            "plans_resolved": 0,
            "plans_unresolved": 0,
            "claim_nodes": 0,
            "category_benefit_links": 0,
            "exclusion_service_links": 0,
            "plan_benefit_lookups": 0,
            "disease_benefit_links": 0,
            "disease_exclusion_links": 0,
            "sign_benefit_links": 0,
            "sign_exclusion_links": 0,
            "benefit_clause_links": 0,
            "clause_reference_rulebook_links": 0,
            "exclusion_rulebook_links": 0,
        }

    @staticmethod
    def _stable_id(prefix: str, *parts: object) -> str:
        payload = "||".join(str(part or "").strip() for part in parts)
        digest = hashlib.sha1(payload.encode("utf-8")).hexdigest()[:16]
        return f"{prefix}-{digest}"

    def close(self):
        if self.driver:
            self.driver.close()

    # ── Bridge A: Plan Resolution ──────────────────────────────────

    def bridge_a_plan_resolution(self) -> None:
        """Load BaoCaoHoSo and create InsuredPerson → ContractPlan links."""
        logger.info("Bridge A: Plan Resolution from BaoCaoHoSo...")

        if not BAOCAOHOSO_PATH.exists():
            logger.warning(f"BaoCaoHoSo not found: {BAOCAOHOSO_PATH}")
            return

        wb = openpyxl.load_workbook(str(BAOCAOHOSO_PATH), read_only=True)
        ws = wb[wb.sheetnames[0]]

        # Also load benefit pack for plan limit lookups
        benefit_limits = self._load_benefit_limits()

        with self.driver.session() as session:
            # Ensure indexes
            session.run(
                "CREATE INDEX claim_ho_so_idx IF NOT EXISTS "
                "FOR (c:Claim) ON (c.ho_so_id)"
            )

            rows_processed = 0
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i == 0:
                    continue  # skip header

                vals = list(row)
                ho_so_id = str(vals[1]) if vals[1] else None
                if not ho_so_id:
                    continue

                insurer_raw = str(vals[0]) if vals[0] else ""
                contract_num = str(vals[2]) if vals[2] else ""
                company_name = str(vals[3]) if vals[3] else ""
                plan_raw = str(vals[31]) if vals[31] else ""
                claim_type = str(vals[13]) if vals[13] else ""
                diagnosis = str(vals[19]) if vals[19] else ""
                csyt = str(vals[18]) if vals[18] else ""
                rejected = str(vals[43]).lower() == "true" if vals[43] else False

                # Parse amounts (Vietnamese format: 1.042.300)
                req_raw = str(vals[11]) if vals[11] else "0"
                paid_raw = str(vals[12]) if vals[12] else "0"
                try:
                    requested = float(req_raw.replace(".", "").replace(",", "."))
                except (ValueError, AttributeError):
                    requested = 0.0
                try:
                    paid = float(paid_raw.replace(".", "").replace(",", "."))
                except (ValueError, AttributeError):
                    paid = 0.0

                insurer_code = normalize_insurer(insurer_raw)
                plan_resolved = resolve_plan_name(plan_raw)

                # Extract ICD code from diagnosis
                icd_match = re.match(r"([A-Z]\d{2}(?:\.\d{1,2})?)", diagnosis)
                icd_code = icd_match.group(1) if icd_match else ""

                # Create Claim node with enriched data
                session.run("""
                    MERGE (cl:Claim {ho_so_id: $ho_so_id})
                    SET cl.namespace = $namespace,
                        cl.insurer_code = $insurer_code,
                        cl.insurer_raw = $insurer_raw,
                        cl.contract_number = $contract_num,
                        cl.company_name = $company_name,
                        cl.plan_raw = $plan_raw,
                        cl.claim_type = $claim_type,
                        cl.diagnosis = $diagnosis,
                        cl.icd_code = $icd_code,
                        cl.csyt = $csyt,
                        cl.requested_vnd = $requested,
                        cl.paid_vnd = $paid,
                        cl.rejected = $rejected,
                        cl.benefit_contract_id = $benefit_contract_id,
                        cl.benefit_plan_name = $benefit_plan_name,
                        cl.source_file = $source_file
                """,
                    ho_so_id=ho_so_id,
                    namespace=INSURANCE_NAMESPACE,
                    insurer_code=insurer_code,
                    insurer_raw=insurer_raw[:60],
                    contract_num=contract_num,
                    company_name=company_name[:60],
                    plan_raw=plan_raw[:80],
                    claim_type=claim_type,
                    diagnosis=diagnosis[:120],
                    icd_code=icd_code,
                    csyt=csyt[:60],
                    requested=requested,
                    paid=paid,
                    rejected=rejected,
                    benefit_contract_id=plan_resolved["contract_id"] if plan_resolved else "",
                    benefit_plan_name=plan_resolved["plan_name"] if plan_resolved else "",
                    source_file=str(BAOCAOHOSO_PATH.relative_to(_CLAIMS_ROOT)),
                )

                # Link Claim → ContractPlan (if resolved)
                if plan_resolved:
                    plan_id = f"{plan_resolved['contract_id']}:{plan_resolved['plan_name']}"
                    session.run("""
                        MATCH (cl:Claim {ho_so_id: $ho_so_id})
                        MATCH (p:ContractPlan {plan_id: $plan_id})
                        MERGE (cl)-[r:UNDER_PLAN]->(p)
                        SET r.namespace = $namespace
                    """, ho_so_id=ho_so_id, plan_id=plan_id, namespace=INSURANCE_NAMESPACE)

                    # Also add benefit limits from plan
                    limits = benefit_limits.get(plan_id, {})
                    if limits:
                        session.run("""
                            MATCH (cl:Claim {ho_so_id: $ho_so_id})
                            SET cl.plan_limit_per_visit = $per_visit,
                                cl.plan_limit_per_year = $per_year
                        """,
                            ho_so_id=ho_so_id,
                            per_visit=limits.get("per_visit", 0),
                            per_year=limits.get("per_year", 0),
                        )
                        self.stats["plan_benefit_lookups"] += 1

                    self.stats["plans_resolved"] += 1
                else:
                    self.stats["plans_unresolved"] += 1

                # Link Claim → Insurer
                session.run("""
                    MATCH (cl:Claim {ho_so_id: $ho_so_id})
                    MERGE (i:Insurer {name: $insurer_code})
                    SET i.namespace = coalesce(i.namespace, $namespace)
                    MERGE (cl)-[r:INSURED_BY]->(i)
                    SET r.namespace = $namespace
                """, ho_so_id=ho_so_id, insurer_code=insurer_code, namespace=INSURANCE_NAMESPACE)

                rows_processed += 1
                self.stats["claim_nodes"] += 1

        wb.close()
        logger.info(
            f"  Bridge A done: {rows_processed} claims, "
            f"{self.stats['plans_resolved']} resolved, "
            f"{self.stats['plans_unresolved']} unresolved, "
            f"{self.stats['plan_benefit_lookups']} with benefit limits"
        )

    def _load_benefit_limits(self) -> dict[str, dict]:
        """Extract per-visit and per-year limits from benefit pack by plan."""
        if not BENEFIT_PACK_PATH.exists():
            return {}

        with open(BENEFIT_PACK_PATH, "r", encoding="utf-8") as f:
            bp = json.load(f)

        limits: dict[str, dict] = {}
        for contract in bp.get("contract_catalog", {}).get("contracts", []):
            contract_id = contract.get("contract_id", "")
            for entry in contract.get("benefit_entries", []):
                label = entry.get("entry_label", "").lower()
                coverage = entry.get("coverage_by_plan", {})

                for plan_name, value in coverage.items():
                    plan_id = f"{contract_id}:{plan_name}"
                    if plan_id not in limits:
                        limits[plan_id] = {}

                    if not value:
                        continue

                    # Parse numeric limits
                    try:
                        num_val = float(str(value).replace(".", "").replace(",", "."))
                    except (ValueError, TypeError):
                        continue

                    # Detect limit type from label
                    if "giới hạn" in label and ("lần" in label or "khám" in label):
                        limits[plan_id]["per_visit"] = num_val
                    elif "giới hạn" in label and ("năm" in label or "người" in label):
                        if num_val > limits[plan_id].get("per_year", 0):
                            limits[plan_id]["per_year"] = num_val

        return limits

    # ── Bridge B: Category → Benefit ───────────────────────────────

    def bridge_b_category_benefit(self) -> None:
        """Derive contract-specific benefit support from evidence-backed interpretation links."""
        logger.info("Bridge B: Evidence-backed service → contract benefit mapping...")

        with self.driver.session() as session:
            rows = session.run("""
                MATCH (svc:CIService)-[support:SUPPORTED_BY_BENEFIT]->(bi:BenefitInterpretation)<-[:INTERPRETED_AS]-(b:Benefit)
                WHERE coalesce(support.namespace, '') = $namespace
                   OR coalesce(b.namespace, '') = $namespace
                RETURN svc.service_code AS service_code,
                       svc.service_name AS service_name,
                       b.entry_id AS benefit_entry_id,
                       b.contract_id AS contract_id,
                       b.entry_label AS benefit_label,
                       bi.entry_id AS interpretation_entry_id,
                       bi.canonical_name AS interpretation_name,
                       support.support_rows AS support_rows,
                       support.claim_count AS claim_count,
                       support.total_detail_amount_vnd AS total_detail_amount_vnd,
                       support.status_distribution_json AS status_distribution_json
            """, namespace=INSURANCE_NAMESPACE).data()

            for row in rows:
                session.run("""
                    MATCH (svc:CIService {service_code: $service_code})
                    MATCH (b:Benefit {entry_id: $benefit_entry_id})
                    MERGE (svc)-[r:SUPPORTED_BY_CONTRACT_BENEFIT]->(b)
                    SET r.namespace = $namespace,
                        r.contract_id = $contract_id,
                        r.interpretation_entry_id = $interpretation_entry_id,
                        r.interpretation_name = $interpretation_name,
                        r.support_rows = $support_rows,
                        r.claim_count = $claim_count,
                        r.total_detail_amount_vnd = $total_detail_amount_vnd,
                        r.status_distribution_json = $status_distribution_json,
                        r.source = 'benefit_detail_evidence'
                """,
                    service_code=row.get("service_code", ""),
                    benefit_entry_id=row.get("benefit_entry_id", ""),
                    namespace=INSURANCE_NAMESPACE,
                    contract_id=row.get("contract_id", ""),
                    interpretation_entry_id=row.get("interpretation_entry_id", ""),
                    interpretation_name=row.get("interpretation_name", ""),
                    support_rows=int(row.get("support_rows", 0) or 0),
                    claim_count=int(row.get("claim_count", 0) or 0),
                    total_detail_amount_vnd=int(row.get("total_detail_amount_vnd", 0) or 0),
                    status_distribution_json=row.get("status_distribution_json", "{}"),
                )
                self.stats["category_benefit_links"] += 1

        logger.info(
            "  Bridge B done: %s evidence-backed service→benefit links",
            self.stats["category_benefit_links"],
        )

    # ── Bridge C: Historical Exclusion → CIService ─────────────────

    def bridge_c_exclusion_service(self) -> None:
        """Link CIService to exclusion patterns using historical claim data."""
        logger.info("Bridge C: Historical Exclusion → CIService...")

        if not EXCLUSION_SIGNAL_PATH.exists():
            logger.warning(f"Exclusion signal catalog not found: {EXCLUSION_SIGNAL_PATH}")
            return

        with open(EXCLUSION_SIGNAL_PATH, "r", encoding="utf-8") as f:
            catalog = json.load(f)

        # Also load exclusion pack for reason→code mapping
        reason_to_code = {}
        if EXCLUSION_PACK_PATH.exists():
            with open(EXCLUSION_PACK_PATH, "r", encoding="utf-8") as f:
                ep = json.load(f)
            for usage in ep.get("reason_usage", []):
                reason_text = usage.get("reason", "")
                code = usage.get("code", "")
                if reason_text and code:
                    reason_to_code[reason_text] = code

        with self.driver.session() as session:
            # Ensure ExclusionPattern nodes exist
            session.run(
                "CREATE INDEX exc_pattern_reason_idx IF NOT EXISTS "
                "FOR (ep:ExclusionPattern) ON (ep.reason)"
            )

            contracts = catalog.get("contracts", [])
            total_links = 0

            for contract in contracts:
                contract_name = contract.get("contract_name", "")
                insurer = contract.get("insurer", "")

                for reason_profile in contract.get("reason_profiles", []):
                    atomic_reason = reason_profile.get("atomic_reason", "")
                    reason_rows = reason_profile.get("rows", 0)

                    if not atomic_reason:
                        continue

                    # Map reason text to exclusion code
                    exc_code = reason_to_code.get(atomic_reason, "")

                    # Create or update ExclusionPattern node
                    pattern_id = self._stable_id("EXCPAT", contract_name, insurer, atomic_reason)
                    session.run("""
                        MERGE (ep:ExclusionPattern {pattern_id: $pattern_id})
                        SET ep.namespace = $namespace,
                            ep.reason = $reason,
                            ep.exclusion_code = $exc_code,
                            ep.contract_name = $contract_name,
                            ep.insurer = $insurer,
                            ep.source_file = $source_file,
                            ep.total_rows = coalesce(ep.total_rows, 0) + $rows
                    """,
                        pattern_id=pattern_id,
                        namespace=INSURANCE_NAMESPACE,
                        reason=atomic_reason,
                        exc_code=exc_code,
                        contract_name=contract_name,
                        insurer=insurer,
                        source_file=str(EXCLUSION_SIGNAL_PATH.relative_to(_CLAIMS_ROOT)),
                        rows=reason_rows,
                    )

                    # Link ExclusionPattern → Exclusion (if code exists)
                    if exc_code:
                        session.run("""
                            MATCH (ep:ExclusionPattern {pattern_id: $pattern_id})
                            MATCH (e:Exclusion {code: $code})
                            WHERE coalesce(e.namespace, '') = $namespace
                            MERGE (ep)-[r:MATCHES_EXCLUSION]->(e)
                            SET r.namespace = $namespace
                        """, pattern_id=pattern_id, code=exc_code, namespace=INSURANCE_NAMESPACE)

                    # Link CIService → ExclusionPattern for top services
                    top_services = reason_profile.get("top_services", [])
                    for svc in top_services:
                        service_code = svc.get("service_code", "")
                        svc_rows = svc.get("rows", 0)

                        if not service_code or svc_rows < 2:
                            continue

                        session.run("""
                            MATCH (ci:CIService {service_code: $service_code})
                            MATCH (ep:ExclusionPattern {pattern_id: $pattern_id})
                            MERGE (ci)-[r:HISTORICALLY_EXCLUDED]->(ep)
                            ON CREATE SET r.rows = $rows,
                                          r.contract_name = $contract_name,
                                          r.insurer = $insurer,
                                          r.reason = $reason,
                                          r.namespace = $namespace,
                                          r.source = 'contract_clause_service_catalog'
                            ON MATCH SET r.rows = r.rows + $rows,
                                          r.reason = $reason,
                                          r.namespace = $namespace
                        """,
                            service_code=service_code,
                            pattern_id=pattern_id,
                            rows=svc_rows,
                            contract_name=contract_name,
                            insurer=insurer,
                            reason=atomic_reason,
                            namespace=INSURANCE_NAMESPACE,
                        )
                        total_links += 1

            self.stats["exclusion_service_links"] = total_links

        logger.info(
            f"  Bridge C done: {total_links} service→exclusion pattern links"
        )

    # ── Bridge D: Medical context → contract benefit/exclusion ──────────────

    def bridge_d_medical_context(self) -> None:
        """Project disease/sign context into contract benefit and exclusion graph."""
        logger.info("Bridge D: Medical context → insurance graph...")

        with self.driver.session() as session:
            disease_support_rows = session.run(
                """
                MATCH (d:CIDisease {namespace:$claims_ns})-[ds:CI_INDICATES_SERVICE]->(svc:CIService)
                MATCH (svc)-[sb:SUPPORTED_BY_CONTRACT_BENEFIT]->(b:Benefit {namespace:$ins_ns})
                RETURN d.disease_id AS disease_id,
                       d.disease_name AS disease_name,
                       svc.service_code AS service_code,
                       svc.service_name AS service_name,
                       ds.roles_json AS roles_json,
                       coalesce(sb.contract_id, b.contract_id) AS contract_id,
                       b.entry_id AS benefit_entry_id,
                       b.entry_label AS benefit_label,
                       b.canonical_name AS benefit_canonical_name,
                       coalesce(sb.support_rows, 0) AS support_rows,
                       coalesce(sb.claim_count, 0) AS claim_count
                """,
                claims_ns="claims_insights_explorer_v1",
                ins_ns=INSURANCE_NAMESPACE,
            ).data()

            disease_support_agg: dict[tuple[str, str, str], dict[str, Any]] = {}
            for row in disease_support_rows:
                disease_id = row.get("disease_id", "")
                contract_id = row.get("contract_id", "")
                benefit_entry_id = row.get("benefit_entry_id", "")
                if not disease_id or not contract_id or not benefit_entry_id:
                    continue
                key = (disease_id, contract_id, benefit_entry_id)
                bucket = disease_support_agg.setdefault(
                    key,
                    {
                        "disease_name": row.get("disease_name", ""),
                        "benefit_label": row.get("benefit_label", ""),
                        "benefit_canonical_name": row.get("benefit_canonical_name", ""),
                        "service_codes": set(),
                        "service_names": set(),
                        "support_rows": 0,
                        "claim_count": 0,
                        "role_counter": Counter(),
                    },
                )
                bucket["service_codes"].add(row.get("service_code", ""))
                bucket["service_names"].add(row.get("service_name", ""))
                bucket["support_rows"] += int(row.get("support_rows", 0) or 0)
                bucket["claim_count"] += int(row.get("claim_count", 0) or 0)
                for role in json.loads(row.get("roles_json") or "[]"):
                    role_text = _text(role)
                    if role_text:
                        bucket["role_counter"][role_text] += 1

            for (disease_id, contract_id, benefit_entry_id), bucket in disease_support_agg.items():
                session.run(
                    """
                    MATCH (d:CIDisease {disease_id:$disease_id, namespace:$claims_ns})
                    MATCH (b:Benefit {entry_id:$benefit_entry_id, namespace:$ins_ns})
                    MERGE (d)-[r:CI_SUPPORTS_CONTRACT_BENEFIT {contract_id:$contract_id, benefit_entry_id:$benefit_entry_id}]->(b)
                    SET r.namespace = $ins_ns,
                        r.source = 'medical_context_projection',
                        r.supporting_service_count = $supporting_service_count,
                        r.support_rows = $support_rows,
                        r.claim_count = $claim_count,
                        r.service_codes_json = $service_codes_json,
                        r.service_names_json = $service_names_json,
                        r.medical_role_distribution_json = $role_distribution_json
                    """,
                    disease_id=disease_id,
                    benefit_entry_id=benefit_entry_id,
                    contract_id=contract_id,
                    claims_ns="claims_insights_explorer_v1",
                    ins_ns=INSURANCE_NAMESPACE,
                    supporting_service_count=len(bucket["service_codes"]),
                    support_rows=bucket["support_rows"],
                    claim_count=bucket["claim_count"],
                    service_codes_json=json.dumps(sorted(code for code in bucket["service_codes"] if code), ensure_ascii=False),
                    service_names_json=json.dumps(sorted(name for name in bucket["service_names"] if name), ensure_ascii=False),
                    role_distribution_json=json.dumps(dict(bucket["role_counter"]), ensure_ascii=False),
                )
                self.stats["disease_benefit_links"] += 1

            sign_support_rows = session.run(
                """
                MATCH (d:CIDisease {namespace:$claims_ns})-[:CI_HAS_SIGN]->(sg:CISign)
                MATCH (d)-[ds:CI_INDICATES_SERVICE]->(svc:CIService)
                MATCH (svc)-[sb:SUPPORTED_BY_CONTRACT_BENEFIT]->(b:Benefit {namespace:$ins_ns})
                RETURN sg.sign_id AS sign_id,
                       sg.text AS sign_label,
                       d.disease_id AS disease_id,
                       d.disease_name AS disease_name,
                       svc.service_code AS service_code,
                       svc.service_name AS service_name,
                       ds.roles_json AS roles_json,
                       coalesce(sb.contract_id, b.contract_id) AS contract_id,
                       b.entry_id AS benefit_entry_id,
                       b.entry_label AS benefit_label,
                       coalesce(sb.support_rows, 0) AS support_rows
                """,
                claims_ns="claims_insights_explorer_v1",
                ins_ns=INSURANCE_NAMESPACE,
            ).data()

            sign_support_agg: dict[tuple[str, str, str], dict[str, Any]] = {}
            for row in sign_support_rows:
                sign_id = row.get("sign_id", "")
                contract_id = row.get("contract_id", "")
                benefit_entry_id = row.get("benefit_entry_id", "")
                if not sign_id or not contract_id or not benefit_entry_id:
                    continue
                key = (sign_id, contract_id, benefit_entry_id)
                bucket = sign_support_agg.setdefault(
                    key,
                    {
                        "sign_label": row.get("sign_label", ""),
                        "service_codes": set(),
                        "disease_ids": set(),
                        "disease_names": set(),
                        "support_rows": 0,
                        "role_counter": Counter(),
                    },
                )
                bucket["service_codes"].add(row.get("service_code", ""))
                bucket["disease_ids"].add(row.get("disease_id", ""))
                bucket["disease_names"].add(row.get("disease_name", ""))
                bucket["support_rows"] += int(row.get("support_rows", 0) or 0)
                for role in json.loads(row.get("roles_json") or "[]"):
                    role_text = _text(role)
                    if role_text:
                        bucket["role_counter"][role_text] += 1

            for (sign_id, contract_id, benefit_entry_id), bucket in sign_support_agg.items():
                session.run(
                    """
                    MATCH (sg:CISign {sign_id:$sign_id})
                    MATCH (b:Benefit {entry_id:$benefit_entry_id, namespace:$ins_ns})
                    MERGE (sg)-[r:CI_SUPPORTS_CONTRACT_BENEFIT {contract_id:$contract_id, benefit_entry_id:$benefit_entry_id}]->(b)
                    SET r.namespace = $ins_ns,
                        r.source = 'medical_context_projection',
                        r.supporting_service_count = $supporting_service_count,
                        r.supporting_disease_count = $supporting_disease_count,
                        r.support_rows = $support_rows,
                        r.service_codes_json = $service_codes_json,
                        r.disease_ids_json = $disease_ids_json,
                        r.disease_names_json = $disease_names_json,
                        r.medical_role_distribution_json = $role_distribution_json
                    """,
                    sign_id=sign_id,
                    benefit_entry_id=benefit_entry_id,
                    contract_id=contract_id,
                    ins_ns=INSURANCE_NAMESPACE,
                    supporting_service_count=len(bucket["service_codes"]),
                    supporting_disease_count=len(bucket["disease_ids"]),
                    support_rows=bucket["support_rows"],
                    service_codes_json=json.dumps(sorted(code for code in bucket["service_codes"] if code), ensure_ascii=False),
                    disease_ids_json=json.dumps(sorted(item for item in bucket["disease_ids"] if item), ensure_ascii=False),
                    disease_names_json=json.dumps(sorted(item for item in bucket["disease_names"] if item), ensure_ascii=False),
                    role_distribution_json=json.dumps(dict(bucket["role_counter"]), ensure_ascii=False),
                )
                self.stats["sign_benefit_links"] += 1

            disease_exclusion_rows = session.run(
                """
                MATCH (d:CIDisease {namespace:$claims_ns})-[ds:CI_INDICATES_SERVICE]->(svc:CIService)
                MATCH (svc)-[ex:EXCLUDED_BY_CONTRACT]->(e:Exclusion {namespace:$ins_ns})
                RETURN d.disease_id AS disease_id,
                       d.disease_name AS disease_name,
                       svc.service_code AS service_code,
                       svc.service_name AS service_name,
                       ds.roles_json AS roles_json,
                       ex.contract_id AS contract_id,
                       e.code AS exclusion_code,
                       e.group AS exclusion_group,
                       coalesce(e.reason, '') AS exclusion_reason,
                       coalesce(ex.rows, 0) AS matched_rows
                """,
                claims_ns="claims_insights_explorer_v1",
                ins_ns=INSURANCE_NAMESPACE,
            ).data()

            disease_exclusion_agg: dict[tuple[str, str, str], dict[str, Any]] = {}
            for row in disease_exclusion_rows:
                disease_id = row.get("disease_id", "")
                contract_id = row.get("contract_id", "")
                exclusion_code = row.get("exclusion_code", "")
                if not disease_id or not contract_id or not exclusion_code:
                    continue
                key = (disease_id, contract_id, exclusion_code)
                bucket = disease_exclusion_agg.setdefault(
                    key,
                    {
                        "disease_name": row.get("disease_name", ""),
                        "service_codes": set(),
                        "matched_rows": 0,
                        "role_counter": Counter(),
                        "exclusion_group": row.get("exclusion_group", ""),
                        "exclusion_reason": row.get("exclusion_reason", ""),
                    },
                )
                bucket["service_codes"].add(row.get("service_code", ""))
                bucket["matched_rows"] += int(row.get("matched_rows", 0) or 0)
                for role in json.loads(row.get("roles_json") or "[]"):
                    role_text = _text(role)
                    if role_text:
                        bucket["role_counter"][role_text] += 1

            for (disease_id, contract_id, exclusion_code), bucket in disease_exclusion_agg.items():
                session.run(
                    """
                    MATCH (d:CIDisease {disease_id:$disease_id, namespace:$claims_ns})
                    MATCH (e:Exclusion {code:$exclusion_code, namespace:$ins_ns})
                    MERGE (d)-[r:CI_FLAGGED_BY_CONTRACT_EXCLUSION {contract_id:$contract_id, exclusion_code:$exclusion_code}]->(e)
                    SET r.namespace = $ins_ns,
                        r.source = 'medical_context_projection',
                        r.supporting_service_count = $supporting_service_count,
                        r.matched_rows = $matched_rows,
                        r.service_codes_json = $service_codes_json,
                        r.medical_role_distribution_json = $role_distribution_json,
                        r.resolution_type = $resolution_type
                    """,
                    disease_id=disease_id,
                    exclusion_code=exclusion_code,
                    contract_id=contract_id,
                    claims_ns="claims_insights_explorer_v1",
                    ins_ns=INSURANCE_NAMESPACE,
                    supporting_service_count=len(bucket["service_codes"]),
                    matched_rows=bucket["matched_rows"],
                    service_codes_json=json.dumps(sorted(code for code in bucket["service_codes"] if code), ensure_ascii=False),
                    role_distribution_json=json.dumps(dict(bucket["role_counter"]), ensure_ascii=False),
                    resolution_type=_classify_exclusion_resolution(
                        exclusion_code,
                        bucket.get("exclusion_group", ""),
                        bucket.get("exclusion_reason", ""),
                    ),
                )
                self.stats["disease_exclusion_links"] += 1

            sign_exclusion_rows = session.run(
                """
                MATCH (d:CIDisease {namespace:$claims_ns})-[:CI_HAS_SIGN]->(sg:CISign)
                MATCH (d)-[ds:CI_INDICATES_SERVICE]->(svc:CIService)
                MATCH (svc)-[ex:EXCLUDED_BY_CONTRACT]->(e:Exclusion {namespace:$ins_ns})
                RETURN sg.sign_id AS sign_id,
                       sg.text AS sign_label,
                       d.disease_id AS disease_id,
                       svc.service_code AS service_code,
                       ds.roles_json AS roles_json,
                       ex.contract_id AS contract_id,
                       e.code AS exclusion_code,
                       e.group AS exclusion_group,
                       coalesce(e.reason, '') AS exclusion_reason,
                       coalesce(ex.rows, 0) AS matched_rows
                """,
                claims_ns="claims_insights_explorer_v1",
                ins_ns=INSURANCE_NAMESPACE,
            ).data()

            sign_exclusion_agg: dict[tuple[str, str, str], dict[str, Any]] = {}
            for row in sign_exclusion_rows:
                sign_id = row.get("sign_id", "")
                contract_id = row.get("contract_id", "")
                exclusion_code = row.get("exclusion_code", "")
                if not sign_id or not contract_id or not exclusion_code:
                    continue
                key = (sign_id, contract_id, exclusion_code)
                bucket = sign_exclusion_agg.setdefault(
                    key,
                    {
                        "service_codes": set(),
                        "disease_ids": set(),
                        "matched_rows": 0,
                        "role_counter": Counter(),
                        "exclusion_group": row.get("exclusion_group", ""),
                        "exclusion_reason": row.get("exclusion_reason", ""),
                    },
                )
                bucket["service_codes"].add(row.get("service_code", ""))
                bucket["disease_ids"].add(row.get("disease_id", ""))
                bucket["matched_rows"] += int(row.get("matched_rows", 0) or 0)
                for role in json.loads(row.get("roles_json") or "[]"):
                    role_text = _text(role)
                    if role_text:
                        bucket["role_counter"][role_text] += 1

            for (sign_id, contract_id, exclusion_code), bucket in sign_exclusion_agg.items():
                session.run(
                    """
                    MATCH (sg:CISign {sign_id:$sign_id})
                    MATCH (e:Exclusion {code:$exclusion_code, namespace:$ins_ns})
                    MERGE (sg)-[r:CI_FLAGGED_BY_CONTRACT_EXCLUSION {contract_id:$contract_id, exclusion_code:$exclusion_code}]->(e)
                    SET r.namespace = $ins_ns,
                        r.source = 'medical_context_projection',
                        r.supporting_service_count = $supporting_service_count,
                        r.supporting_disease_count = $supporting_disease_count,
                        r.matched_rows = $matched_rows,
                        r.service_codes_json = $service_codes_json,
                        r.disease_ids_json = $disease_ids_json,
                        r.medical_role_distribution_json = $role_distribution_json,
                        r.resolution_type = $resolution_type
                    """,
                    sign_id=sign_id,
                    exclusion_code=exclusion_code,
                    contract_id=contract_id,
                    ins_ns=INSURANCE_NAMESPACE,
                    supporting_service_count=len(bucket["service_codes"]),
                    supporting_disease_count=len(bucket["disease_ids"]),
                    matched_rows=bucket["matched_rows"],
                    service_codes_json=json.dumps(sorted(code for code in bucket["service_codes"] if code), ensure_ascii=False),
                    disease_ids_json=json.dumps(sorted(item for item in bucket["disease_ids"] if item), ensure_ascii=False),
                    role_distribution_json=json.dumps(dict(bucket["role_counter"]), ensure_ascii=False),
                    resolution_type=_classify_exclusion_resolution(
                        exclusion_code,
                        bucket.get("exclusion_group", ""),
                        bucket.get("exclusion_reason", ""),
                    ),
                )
                self.stats["sign_exclusion_links"] += 1

        logger.info(
            "  Bridge D done: disease-benefit=%s, disease-exclusion=%s, sign-benefit=%s, sign-exclusion=%s",
            self.stats["disease_benefit_links"],
            self.stats["disease_exclusion_links"],
            self.stats["sign_benefit_links"],
            self.stats["sign_exclusion_links"],
        )

    # ── Bridge E: Policy grounding ──────────────────────────────────────────

    def bridge_e_policy_grounding(self) -> None:
        """Ground benefits/exclusions into clauses and rulebooks for explainable decisions."""
        logger.info("Bridge E: Policy grounding...")

        with self.driver.session() as session:
            benefit_rows = session.run(
                """
                MATCH (b:Benefit {namespace:$ns})
                OPTIONAL MATCH (b)-[:INTERPRETED_AS]->(bi:BenefitInterpretation {namespace:$ns})
                RETURN b.entry_id AS entry_id,
                       b.contract_id AS contract_id,
                       b.entry_label AS entry_label,
                       b.canonical_name AS canonical_name,
                       b.major_section AS major_section,
                       collect(DISTINCT bi.canonical_name) AS interpretation_names,
                       collect(DISTINCT bi.definition_text) AS definition_texts,
                       collect(DISTINCT bi.interpretation_text) AS interpretation_texts
                """,
                ns=INSURANCE_NAMESPACE,
            ).data()
            clause_rows = session.run(
                """
                MATCH (cc:ContractClause {namespace:$ns})
                RETURN cc.clause_id AS clause_id,
                       cc.contract_id AS contract_id,
                       cc.sheet_name AS sheet_name,
                       cc.section AS section,
                       cc.clause_code AS clause_code,
                       cc.clause_title AS clause_title,
                       cc.clause_body AS clause_body
                """,
                ns=INSURANCE_NAMESPACE,
            ).data()
            clause_by_contract: dict[str, list[dict[str, Any]]] = {}
            for row in clause_rows:
                clause_by_contract.setdefault(_text(row.get("contract_id")), []).append(row)

            for benefit in benefit_rows:
                contract_id = _text(benefit.get("contract_id"))
                entry_id = _text(benefit.get("entry_id"))
                if not contract_id or not entry_id:
                    continue
                if _is_generic_benefit(
                    _text(benefit.get("entry_label")),
                    _text(benefit.get("canonical_name")),
                ):
                    continue
                source_phrases = [
                    _text(benefit.get("entry_label")),
                    _text(benefit.get("canonical_name")),
                ]
                source_phrases.extend(_text(item) for item in (benefit.get("interpretation_names") or [])[:2])
                source_phrases = [item for item in source_phrases if item]
                source_text = " ".join(source_phrases)
                candidates: list[tuple[float, list[str], dict[str, Any]]] = []
                for clause in clause_by_contract.get(contract_id, []):
                    clause_text = " ".join(
                        part for part in [
                            _text(clause.get("clause_title")),
                            _text(clause.get("clause_body")),
                            _text(clause.get("section")),
                        ] if part
                    )
                    clause_folded = _ascii_fold(clause_text)
                    score = 0.0
                    matched_terms: list[str] = []
                    for phrase in source_phrases:
                        phrase_folded = _ascii_fold(phrase)
                        if len(phrase_folded) >= 12 and phrase_folded in clause_folded:
                            score = 0.98
                            matched_terms = [phrase_folded[:48]]
                            break
                    if score < 0.98:
                        score, matched_terms = _score_overlap(source_text, clause_text)
                    if score < 0.75 or len(matched_terms) < 2:
                        continue
                    candidates.append((score, matched_terms, clause))

                candidates.sort(key=lambda item: (-item[0], -(len(item[1])), _text(item[2].get("clause_id"))))
                for score, matched_terms, clause in candidates[:2]:
                    session.run(
                        """
                        MATCH (b:Benefit {entry_id:$entry_id, namespace:$ns})
                        MATCH (cc:ContractClause {clause_id:$clause_id, namespace:$ns})
                        MERGE (b)-[r:GROUNDED_IN_CONTRACT_CLAUSE]->(cc)
                        SET r.namespace = $ns,
                            r.contract_id = $contract_id,
                            r.match_score = $match_score,
                            r.match_method = 'text_overlap',
                            r.matched_terms = $matched_terms,
                            r.source = 'policy_grounding'
                        """,
                        entry_id=entry_id,
                        clause_id=clause.get("clause_id"),
                        ns=INSURANCE_NAMESPACE,
                        contract_id=contract_id,
                        match_score=float(score),
                        matched_terms=matched_terms,
                    )
                    self.stats["benefit_clause_links"] += 1

            rulebook_rows = session.run(
                """
                MATCH (r:Rulebook {namespace:$ns})
                RETURN r.rulebook_id AS rulebook_id,
                       r.rule_code AS rule_code,
                       r.display_name AS display_name,
                       r.insurer AS insurer
                """,
                ns=INSURANCE_NAMESPACE,
            ).data()
            clause_reference_rows = session.run(
                """
                MATCH (cr:ClauseReference {namespace:$ns})
                RETURN cr.clause_id AS clause_id,
                       cr.rule_name AS rule_name,
                       cr.clause_reference AS clause_reference
                """,
                ns=INSURANCE_NAMESPACE,
            ).data()

            for ref in clause_reference_rows:
                rule_name = _ascii_fold(ref.get("rule_name"))
                clause_reference = _ascii_fold(ref.get("clause_reference"))
                if not rule_name and not clause_reference:
                    continue
                matches: list[dict[str, Any]] = []
                for rulebook in rulebook_rows:
                    rule_code = _ascii_fold(rulebook.get("rule_code"))
                    insurer = _ascii_fold(rulebook.get("insurer"))
                    display_name = _ascii_fold(rulebook.get("display_name"))
                    score = 0.0
                    if rule_code and (rule_code in rule_name or rule_code in clause_reference):
                        score += 0.7
                    if insurer and (insurer in rule_name or insurer in clause_reference):
                        score += 0.2
                    if display_name and display_name in clause_reference:
                        score += 0.2
                    if score >= 0.6:
                        matches.append({"rulebook_id": rulebook.get("rulebook_id"), "score": min(score, 1.0)})

                matches.sort(key=lambda item: (-float(item["score"]), _text(item["rulebook_id"])))
                for item in matches[:2]:
                    session.run(
                        """
                        MATCH (cr:ClauseReference {clause_id:$clause_id, namespace:$ns})
                        MATCH (r:Rulebook {rulebook_id:$rulebook_id, namespace:$ns})
                        MERGE (cr)-[rel:REFERS_TO_RULEBOOK]->(r)
                        SET rel.namespace = $ns,
                            rel.match_score = $match_score,
                            rel.source = 'policy_grounding'
                        """,
                        clause_id=ref.get("clause_id"),
                        rulebook_id=item["rulebook_id"],
                        ns=INSURANCE_NAMESPACE,
                        match_score=float(item["score"]),
                    )
                    self.stats["clause_reference_rulebook_links"] += 1

            exclusion_rule_rows = session.run(
                """
                MATCH (e:Exclusion {namespace:$ns})-[:HAS_REASON]->(er:ExclusionReason)<-[:LINKS_TO]-(cr:ClauseReference {namespace:$ns})-[rr:REFERS_TO_RULEBOOK]->(rb:Rulebook {namespace:$ns})
                RETURN e.code AS exclusion_code,
                       rb.rulebook_id AS rulebook_id,
                       count(DISTINCT cr) AS clause_ref_count,
                       avg(coalesce(rr.match_score, 0.0)) AS avg_match_score
                """,
                ns=INSURANCE_NAMESPACE,
            ).data()

            for row in exclusion_rule_rows:
                session.run(
                    """
                    MATCH (e:Exclusion {code:$exclusion_code, namespace:$ns})
                    MATCH (rb:Rulebook {rulebook_id:$rulebook_id, namespace:$ns})
                    MERGE (e)-[r:SUPPORTED_BY_RULEBOOK]->(rb)
                    SET r.namespace = $ns,
                        r.clause_reference_count = $clause_reference_count,
                        r.match_score = $match_score,
                        r.source = 'policy_grounding'
                    """,
                    exclusion_code=row.get("exclusion_code"),
                    rulebook_id=row.get("rulebook_id"),
                    ns=INSURANCE_NAMESPACE,
                    clause_reference_count=int(row.get("clause_ref_count", 0) or 0),
                    match_score=float(row.get("avg_match_score", 0.0) or 0.0),
                )
                self.stats["exclusion_rulebook_links"] += 1

        logger.info(
            "  Bridge E done: benefit-clause=%s, clause-ref-rulebook=%s, exclusion-rulebook=%s",
            self.stats["benefit_clause_links"],
            self.stats["clause_reference_rulebook_links"],
            self.stats["exclusion_rulebook_links"],
        )

    # ── Summary ────────────────────────────────────────────────────

    def print_summary(self) -> None:
        logger.info("=" * 55)
        logger.info("INSURANCE-SERVICE BRIDGE SUMMARY")
        logger.info("=" * 55)
        logger.info(f"Bridge A - Plan Resolution:")
        logger.info(f"  Claims ingested:    {self.stats['claim_nodes']}")
        logger.info(f"  Plans resolved:     {self.stats['plans_resolved']}")
        logger.info(f"  Plans unresolved:   {self.stats['plans_unresolved']}")
        logger.info(f"  Benefit limits set: {self.stats['plan_benefit_lookups']}")
        logger.info(f"Bridge B - Category→Benefit:")
        logger.info(f"  Links created:      {self.stats['category_benefit_links']}")
        logger.info(f"Bridge C - Exclusion→Service:")
        logger.info(f"  Links created:      {self.stats['exclusion_service_links']}")
        logger.info("Bridge D - Medical Context → Insurance:")
        logger.info(f"  Disease→Benefit:    {self.stats['disease_benefit_links']}")
        logger.info(f"  Disease→Exclusion:  {self.stats['disease_exclusion_links']}")
        logger.info(f"  Sign→Benefit:       {self.stats['sign_benefit_links']}")
        logger.info(f"  Sign→Exclusion:     {self.stats['sign_exclusion_links']}")
        logger.info("Bridge E - Policy Grounding:")
        logger.info(f"  Benefit→Clause:     {self.stats['benefit_clause_links']}")
        logger.info(f"  ClauseRef→Rulebook: {self.stats['clause_reference_rulebook_links']}")
        logger.info(f"  Exclusion→Rulebook: {self.stats['exclusion_rulebook_links']}")
        logger.info("=" * 55)


def clear_bridge_data(driver) -> None:
    """Clear bridge-specific data (not core insurance data)."""
    with driver.session() as session:
        logger.info("Clearing bridge data...")
        session.run("""
            MATCH (cl:Claim {namespace: $namespace})-[r:UNDER_PLAN]->()
            DELETE r
        """, namespace=INSURANCE_NAMESPACE)
        session.run("""
            MATCH (cl:Claim {namespace: $namespace})-[r:INSURED_BY]->()
            DELETE r
        """, namespace=INSURANCE_NAMESPACE)
        session.run("""
            MATCH (:CIService)-[r:FALLS_UNDER_BENEFIT]->(b:Benefit)
            WHERE r.namespace = $namespace OR (r.namespace IS NULL AND b.namespace = $namespace)
            DELETE r
        """, namespace=INSURANCE_NAMESPACE)
        session.run("""
            MATCH (:CIService)-[r:SUPPORTED_BY_CONTRACT_BENEFIT]->(b:Benefit)
            WHERE r.namespace = $namespace OR (r.namespace IS NULL AND b.namespace = $namespace)
            DELETE r
        """, namespace=INSURANCE_NAMESPACE)
        session.run("""
            MATCH (:CIService)-[r:HISTORICALLY_EXCLUDED]->(ep:ExclusionPattern)
            WHERE r.namespace = $namespace OR ep.namespace = $namespace
            DELETE r
        """, namespace=INSURANCE_NAMESPACE)
        session.run("""
            MATCH (ep:ExclusionPattern {namespace: $namespace})-[r:MATCHES_EXCLUSION]->()
            DELETE r
        """, namespace=INSURANCE_NAMESPACE)
        session.run("""
            MATCH (:CIDisease {namespace: $claims_namespace})-[r:CI_SUPPORTS_CONTRACT_BENEFIT|CI_FLAGGED_BY_CONTRACT_EXCLUSION]->()
            WHERE r.namespace = $namespace
            DELETE r
        """, namespace=INSURANCE_NAMESPACE, claims_namespace="claims_insights_explorer_v1")
        session.run("""
            MATCH (:CISign)-[r:CI_SUPPORTS_CONTRACT_BENEFIT|CI_FLAGGED_BY_CONTRACT_EXCLUSION]->()
            WHERE r.namespace = $namespace
            DELETE r
        """, namespace=INSURANCE_NAMESPACE)
        session.run("""
            MATCH (:Benefit {namespace: $namespace})-[r:GROUNDED_IN_CONTRACT_CLAUSE]->(:ContractClause {namespace: $namespace})
            DELETE r
        """, namespace=INSURANCE_NAMESPACE)
        session.run("""
            MATCH (:ClauseReference {namespace: $namespace})-[r:REFERS_TO_RULEBOOK]->(:Rulebook {namespace: $namespace})
            DELETE r
        """, namespace=INSURANCE_NAMESPACE)
        session.run("""
            MATCH (:Exclusion {namespace: $namespace})-[r:SUPPORTED_BY_RULEBOOK]->(:Rulebook {namespace: $namespace})
            DELETE r
        """, namespace=INSURANCE_NAMESPACE)
        session.run("MATCH (c:Claim {namespace: $namespace}) DETACH DELETE c", namespace=INSURANCE_NAMESPACE)
        session.run("MATCH (ep:ExclusionPattern {namespace: $namespace}) DETACH DELETE ep", namespace=INSURANCE_NAMESPACE)
        logger.info("  Bridge data cleared")


def main():
    parser = argparse.ArgumentParser(
        description="Bridge insurance and service domains in Neo4j"
    )
    parser.add_argument(
        "--clear", action="store_true",
        help="Clear bridge data before rebuilding",
    )
    parser.add_argument(
        "--bridge", choices=["a", "b", "c", "d", "e", "all"], default="all",
        help="Which bridge to run (default: all)",
    )
    args = parser.parse_args()

    bridge = InsuranceServiceBridge()

    try:
        if args.clear:
            clear_bridge_data(bridge.driver)

        if args.bridge in ("a", "all"):
            bridge.bridge_a_plan_resolution()
        if args.bridge in ("b", "all"):
            bridge.bridge_b_category_benefit()
        if args.bridge in ("c", "all"):
            bridge.bridge_c_exclusion_service()
        if args.bridge in ("d", "all"):
            bridge.bridge_d_medical_context()
        if args.bridge in ("e", "all"):
            bridge.bridge_e_policy_grounding()

        bridge.print_summary()

    except Exception as e:
        logger.error(f"Bridge error: {e}", exc_info=True)
    finally:
        bridge.close()


if __name__ == "__main__":
    main()
