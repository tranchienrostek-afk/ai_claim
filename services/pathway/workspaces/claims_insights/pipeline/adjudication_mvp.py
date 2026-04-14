from __future__ import annotations

import argparse
import json
import unicodedata
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import openpyxl


PROJECT_DIR = Path(__file__).parent.parent
CLAIMS_DIR = PROJECT_DIR / "01_claims"
STANDARDIZE_DIR = PROJECT_DIR / "02_standardize"
ENRICH_DIR = PROJECT_DIR / "03_enrich"
INSURANCE_DIR = PROJECT_DIR / "06_insurance"
OUTPUT_DIR = Path(__file__).parent

MATRIX_PATH = ENRICH_DIR / "service_disease_matrix.json"
CONTRACT_RULES_PATH = INSURANCE_DIR / "contract_rules.json"
CODEBOOK_PATH = STANDARDIZE_DIR / "service_codebook.json"
CLASSIFICATION_LOGS_PATH = CLAIMS_DIR / "classification_logs.json"
BENCHMARK_XLSX_PATH = CLAIMS_DIR / "bao_cao_21_03.xlsx"
DEFAULT_BENCHMARK_OUTPUT = OUTPUT_DIR / "adjudication_benchmark_mvp.json"


def strip_diacritics(text: str) -> str:
    text = str(text or "").lower().replace("đ", "d").replace("Đ", "d")
    nfkd = unicodedata.normalize("NFKD", text)
    return "".join(c for c in nfkd if not unicodedata.combining(c))


def normalize_icd(icd: str | None) -> str:
    raw = str(icd or "").strip().upper()
    if not raw:
        return ""
    return "".join(ch for ch in raw if ch.isalnum() or ch == ".")


def icd_group(icd: str | None) -> str:
    code = normalize_icd(icd)
    if not code:
        return ""
    return code.split(".")[0][:3]


@dataclass
class ClinicalAssessment:
    justified: bool
    confidence: float
    reason: str
    role: str
    linked_conditions: list[str]


@dataclass
class ContractAssessment:
    eligible: bool
    rule_applied: str
    note: str


class AdjudicationMVP:
    def __init__(self) -> None:
        self.matrix_payload = json.loads(MATRIX_PATH.read_text(encoding="utf-8"))
        self.contract_payload = json.loads(CONTRACT_RULES_PATH.read_text(encoding="utf-8"))
        self.codebook_payload = json.loads(CODEBOOK_PATH.read_text(encoding="utf-8"))

        mapper_path = str(STANDARDIZE_DIR)
        import sys

        if mapper_path not in sys.path:
            sys.path.insert(0, mapper_path)
        from service_text_mapper import ServiceTextMapper  # pylint: disable=import-error
        from service_mapping_policy import resolve_mapping_resolution  # pylint: disable=import-error

        self.mapper = ServiceTextMapper(codebook_path=CODEBOOK_PATH)
        self.resolve_mapping_resolution = resolve_mapping_resolution
        self.codebook_by_code = {
            e["service_code"]: e for e in self.codebook_payload.get("codebook", []) if e.get("service_code")
        }

        self.links_by_service: dict[str, list[dict]] = {}
        for link in self.matrix_payload.get("links", []):
            self.links_by_service.setdefault(link["service_code"], []).append(link)

        self.contract_rules = {
            r["contract_id"]: r for r in self.contract_payload.get("contract_rules", [])
        }

    def recognize_service(self, service_name: str) -> dict:
        scored = self.mapper.score_text(service_name, top_k=3)
        suggestions = scored.get("suggestions") or []
        resolution = self.resolve_mapping_resolution(suggestions)
        suggested_code = resolution["suggested_service_code"]
        score = float(resolution["mapper_score"] or 0.0)
        mapping_status = resolution["mapping_resolution"]
        mapping_gap = resolution["mapping_gap"]
        mapping_reason = resolution["mapping_reason"]
        accepted_code = suggested_code if resolution["accepted"] else ""
        entry = self.codebook_by_code.get(accepted_code, {})
        return {
            "service_code": accepted_code,
            "canonical_name": entry.get("canonical_name", service_name),
            "category_code": entry.get("category_code", str((suggestions[0] if suggestions else {}).get("category_code") or "UNKNOWN")),
            "category_name": entry.get("category_name", str((suggestions[0] if suggestions else {}).get("category_name") or "")),
            "mapper_score": score,
            "mapper_confidence": resolution["mapper_confidence"],
            "mapping_status": mapping_status,
            "mapping_gap": mapping_gap,
            "mapping_reason": mapping_reason,
            "suggested_service_code": suggested_code,
            "suggested_canonical_name": resolution["suggested_canonical_name"] or service_name,
            "top_candidates": resolution["top_candidates"],
            "bhyt_price": (entry.get("bhyt") or {}).get("gia_tt39_vnd"),
        }

    def _matrix_best_match(self, service_code: str, icd_code: str) -> dict | None:
        if not service_code:
            return None
        links = self.links_by_service.get(service_code, [])
        if not links:
            return None
        icd_full = normalize_icd(icd_code)
        icd3 = icd_group(icd_code)
        candidates = [
            l
            for l in links
            if l.get("icd10") == icd_full or l.get("icd10_group") == icd3
        ]
        if not candidates:
            return None
        candidates.sort(key=lambda x: (x.get("score", 0.0), x.get("support", {}).get("co_occurrence") or 0), reverse=True)
        return candidates[0]

    def assess_clinical_necessity(
        self,
        service_name: str,
        service_info: dict,
        diagnosis: str,
        icd_code: str,
    ) -> ClinicalAssessment:
        service_code = service_info.get("service_code", "")
        diag_key = strip_diacritics(diagnosis)
        svc_key = strip_diacritics(service_name)

        best = self._matrix_best_match(service_code, icd_code)
        if best:
            linked = [best.get("icd10", "")]
            conf = max(0.6, float(best.get("score", 0.6)))
            return ClinicalAssessment(
                justified=True,
                confidence=min(0.95, conf),
                reason=(
                    f"Matched service-disease matrix: {service_code} ↔ "
                    f"{best.get('icd10')} ({best.get('disease_name')}) via {best.get('evidence')}"
                ),
                role=best.get("role", "diagnostic"),
                linked_conditions=linked,
            )

        # Lightweight safeguard for clearly unrelated pathogen-specific tests.
        pathogen_terms = ["dengue", "ns1", "virus test nhanh", "influenza", "ev71", "rsv", "covid", "sars"]
        viral_diag_terms = ["sot", "virus", "cum", "covid", "dengue", "viem gan virus", "nhiem sieu vi"]
        if any(t in svc_key for t in pathogen_terms) and not any(t in diag_key for t in viral_diag_terms):
            return ClinicalAssessment(
                justified=False,
                confidence=0.82,
                reason="Pathogen-specific test lacks corresponding viral clinical context.",
                role="confirmatory",
                linked_conditions=[],
            )

        # Default: uncertain but potentially reasonable in outpatient workflow.
        return ClinicalAssessment(
            justified=True,
            confidence=0.55,
            reason="No direct matrix evidence; fallback to cautious approve in outpatient diagnostic workflow.",
            role="diagnostic",
            linked_conditions=[],
        )

    def _match_contract(self, contract_id: str | None) -> dict | None:
        if not contract_id:
            return None
        key = contract_id.strip()
        if key in self.contract_rules:
            return self.contract_rules[key]
        key_norm = strip_diacritics(key)
        for name, rule in self.contract_rules.items():
            if strip_diacritics(name) == key_norm:
                return rule
        return None

    def _is_category_covered(self, category_code: str, covered_patterns: list[str]) -> bool:
        if not category_code:
            return False
        for p in covered_patterns:
            if p.endswith("*") and category_code.startswith(p[:-1]):
                return True
            if category_code == p:
                return True
        return False

    def assess_contract_eligibility(
        self,
        contract_id: str | None,
        service_info: dict,
        clinical: ClinicalAssessment,
    ) -> ContractAssessment:
        contract = self._match_contract(contract_id)
        if contract is None:
            return ContractAssessment(
                eligible=True,
                rule_applied="unknown_contract_fallback",
                note="No contract metadata on claim; using medical-necessity fallback.",
            )

        if not self._is_category_covered(service_info.get("category_code", ""), contract.get("covered_categories", [])):
            return ContractAssessment(
                eligible=False,
                rule_applied="category_not_covered",
                note=f"Service category {service_info.get('category_code')} is outside contract coverage.",
            )

        rule_type = contract.get("rule_type", "pay_if_medically_necessary")
        if rule_type == "pay_if_medically_necessary":
            return ContractAssessment(
                eligible=clinical.justified,
                rule_applied=rule_type,
                note="Eligibility follows clinical necessity.",
            )
        if rule_type == "pay_if_positive":
            return ContractAssessment(
                eligible=False,
                rule_applied=rule_type,
                note="Positive test result dependency cannot be evaluated from current claim payload.",
            )
        if rule_type == "pay_if_preauthorized":
            return ContractAssessment(
                eligible=False,
                rule_applied=rule_type,
                note="Pre-authorization evidence is unavailable in current payload.",
            )
        if rule_type == "pay_if_final_icd_match":
            return ContractAssessment(
                eligible=clinical.justified,
                rule_applied=rule_type,
                note="Using matrix/icd proxy match for final ICD rule.",
            )
        return ContractAssessment(
            eligible=clinical.justified,
            rule_applied="fallback_rule",
            note="Fallback contract decision.",
        )

    def detect_anomalies(self, service_name: str, amount: float, service_info: dict, all_services: list[dict]) -> list[str]:
        flags: list[str] = []
        bhyt_price = service_info.get("bhyt_price")
        if bhyt_price and amount:
            ratio = amount / bhyt_price if bhyt_price > 0 else 0
            if ratio > 5:
                flags.append("price_extreme")
            elif ratio > 3:
                flags.append("price_outlier")

        normalized = strip_diacritics(service_name)
        duplicates = sum(1 for s in all_services if strip_diacritics(s.get("service", "")) == normalized)
        if duplicates > 1:
            flags.append("duplicate_service")
        return flags

    def adjudicate_service_line(
        self,
        service_name: str,
        amount: float,
        diagnosis: str,
        icd_code: str,
        contract_id: str | None,
        all_services: list[dict],
    ) -> dict:
        service_info = self.recognize_service(service_name)
        clinical = self.assess_clinical_necessity(service_name, service_info, diagnosis, icd_code)
        contract = self.assess_contract_eligibility(contract_id, service_info, clinical)
        anomalies = self.detect_anomalies(service_name, amount, service_info, all_services)

        if not clinical.justified:
            decision = "deny"
        elif not contract.eligible:
            decision = "deny"
        elif anomalies:
            decision = "approve_with_flag"
        elif clinical.confidence < 0.52:
            decision = "manual_review"
        else:
            decision = "approve"

        confidence = min(
            0.98,
            max(
                0.1,
                clinical.confidence * 0.7
                + (0.2 if contract.eligible else 0.0)
                - (0.08 if anomalies else 0.0),
            ),
        )

        return {
            "service": service_name,
            "standardized_code": service_info.get("service_code"),
            "amount_claimed": amount,
            "amount_reference": service_info.get("bhyt_price"),
            "clinical_assessment": {
                "necessity": "justified" if clinical.justified else "unjustified",
                "reasoning": clinical.reason,
                "service_role": clinical.role,
                "linked_conditions": clinical.linked_conditions,
            },
            "contract_assessment": {
                "eligible": contract.eligible,
                "rule_applied": contract.rule_applied,
                "note": contract.note,
            },
            "decision": decision,
            "confidence": round(confidence, 3),
            "flags": anomalies,
            "explanation": (
                f"Clinical={clinical.justified} ({clinical.reason}); "
                f"Contract={contract.eligible} ({contract.rule_applied}); "
                f"Anomalies={','.join(anomalies) if anomalies else 'none'}."
            ),
        }


def load_claim_contexts() -> dict[str, dict]:
    logs = json.loads(CLASSIFICATION_LOGS_PATH.read_text(encoding="utf-8"))
    contexts = {}
    for rec in logs:
        message_hash = rec.get("message_hash_id")
        inp = rec.get("input", {})
        claim_info = inp.get("claim_info", {})
        if not message_hash:
            continue
        contexts[message_hash] = {
            "diagnosis": claim_info.get("diagnosis", ""),
            "icd": claim_info.get("primary_diagnosis_code"),
            "services": inp.get("claims", []),
            "contract_id": claim_info.get("contract"),
        }
    return contexts


def load_benchmark_rows() -> list[dict]:
    wb = openpyxl.load_workbook(BENCHMARK_XLSX_PATH, data_only=True)
    ws = wb["Feedback CLS"]
    headers = [ws.cell(9, c).value for c in range(1, 13)]
    rows = []
    for r in range(10, ws.max_row + 1):
        row = {headers[i]: ws.cell(r, i + 1).value for i in range(len(headers))}
        if row.get("STT") is None:
            continue
        rows.append(row)
    return rows


def decision_to_payment_status(decision: str) -> str:
    if decision in {"approve", "approve_with_flag"}:
        return "PAYMENT"
    return "REJECT"


def run_benchmark(output_path: Path) -> None:
    engine = AdjudicationMVP()
    contexts = load_claim_contexts()
    rows = load_benchmark_rows()

    correct = 0
    total = 0
    tp = fp = tn = fn = 0
    results = []

    for row in rows:
        total += 1
        message_hash = row.get("message_hash_id")
        ctx = contexts.get(message_hash, {})
        diagnosis = ctx.get("diagnosis", "")
        icd_code = ctx.get("icd") or ""
        services = ctx.get("services", [])
        contract_id = ctx.get("contract_id")

        service = row.get("Nội dung chi phí", "")
        amount = float(row.get("Thành tiền", 0) or 0)
        human = str(row.get("FB paymentStatus") or "").strip().upper()

        out = engine.adjudicate_service_line(
            service_name=service,
            amount=amount,
            diagnosis=diagnosis,
            icd_code=icd_code,
            contract_id=contract_id,
            all_services=services,
        )
        ours = decision_to_payment_status(out["decision"])
        match = ours == human
        if match:
            correct += 1
        if ours == "PAYMENT" and human == "PAYMENT":
            tp += 1
        elif ours == "PAYMENT" and human == "REJECT":
            fp += 1
        elif ours == "REJECT" and human == "REJECT":
            tn += 1
        elif ours == "REJECT" and human == "PAYMENT":
            fn += 1

        results.append(
            {
                "stt": row.get("STT"),
                "service": service,
                "human": human,
                "ours": ours,
                "match": match,
                "decision": out["decision"],
                "confidence": out["confidence"],
                "explanation": out["explanation"],
            }
        )

    acc = round(100.0 * correct / max(total, 1), 2)
    summary = {
        "total": total,
        "correct": correct,
        "accuracy_pct": acc,
        "confusion_matrix": {"tp": tp, "fp": fp, "tn": tn, "fn": fn},
        "precision_pct": round(100.0 * tp / max(tp + fp, 1), 2),
        "recall_pct": round(100.0 * tp / max(tp + fn, 1), 2),
        "results": results,
    }
    output_path.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Benchmark saved: {output_path}")
    print(f"Accuracy: {acc}% ({correct}/{total})")
    print(f"Confusion: TP={tp}, FP={fp}, TN={tn}, FN={fn}")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="AZINSU adjudication MVP")
    parser.add_argument("--benchmark", action="store_true", help="Run 99-line benchmark.")
    parser.add_argument("--output", type=Path, default=DEFAULT_BENCHMARK_OUTPUT)
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    if args.benchmark:
        run_benchmark(args.output)
    else:
        print("Use --benchmark to run MVP benchmark mode.")


if __name__ == "__main__":
    main()
